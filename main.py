#inspectionviewer
import streamlit as st
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from io import BytesIO
from matplotlib import pyplot as plt
import altair as alt
import random
import string
import numpy as np
from PIL import Image, ImageDraw, ImageFont, ImageFilter
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle
from pandas.api.types import is_numeric_dtype, is_datetime64_any_dtype
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
from st_aggrid.shared import JsCode
import pytz
from datetime import datetime, date, timedelta
from urllib.parse import quote
import re
import plotly.express as px
import plotly.graph_objects as go
# =========================================================================
# CONFIG
# =========================================================================
st.set_page_config(
    page_title="S.A.R.A.L Inspection App",
    layout="wide",
    initial_sidebar_state="auto",  # collapsed on mobile by default in many clients
    menu_items={
        "Get Help": "https://wa.me/919022507772",
        "About": "S.A.R.A.L – Safety Abnormality Report & Action List (Solapur Division)",
    },
)

TIMESTAMP_COL_NAME = "Timestamp of Compliance"

# ---- "Inspections" sheet config (used by the WhatsApp report generator tab) ----
_gs_secrets = st.secrets.get("google_sheets", {})
INSPECTIONS_SHEET_ID = _gs_secrets.get("ins_sheet_id")
INSPECTIONS_SHEET_NAME = _gs_secrets.get("ins_sheet_name")
# Field name -> 0-indexed column position (A=0, B=1, C=2, ...) as specified.
INSPECTIONS_COL_MAP = {
    "Name": 1,           # Column B - Inspecting Official's name
    "Phone": 3,           # Column D - phone number
    "Type": 4,            # Column E - Type of Inspection
    "Location": 5,        # Column F - Location
    "Date": 6,             # Column G - Date of Inspection
    "Deficiency": 9,        # Column J - Deficiency noted
    "InspectionBy": 10,      # Column K - Inspection By / Designation & HQ
    "ActionBy": 11,           # Column L - Action By
}

# Characters used for the CAPTCHA text. Visually-confusable characters
# (0/O, 1/I/l) are excluded so a genuine human isn't penalised for a
# reasonable misread.
CAPTCHA_CHARS = "".join(c for c in (string.ascii_uppercase + string.digits) if c not in "0O1IL")
CAPTCHA_LENGTH = 5
CAPTCHA_MAX_ATTEMPTS = 5  # lock out after this many wrong CAPTCHA attempts in a session

# =========================================================================
# GLOBAL RESPONSIVE CSS (mobile-first + adaptive)
# =========================================================================
st.markdown(
    """
<style>
/* ---------- Base / Reset ---------- */
html, body, [class*="css"] {
    font-family: 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif !important;
    -webkit-text-size-adjust: 100%;
    text-size-adjust: 100%;
}

/* Make main content breathe on all devices */
.main .block-container {
    padding-top: 1rem !important;
    padding-bottom: 2rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}

/* Touch-friendly buttons & inputs */
.stButton > button,
.stDownloadButton > button,
div[data-testid="stForm"] button {
    min-height: 44px !important;          /* Apple HIG / Material touch target */
    border-radius: 10px !important;
    font-weight: 600 !important;
    white-space: normal !important;
    word-break: break-word !important;
}

.stTextInput input,
.stTextArea textarea,
.stSelectbox > div,
.stMultiSelect > div {
    min-height: 42px !important;
    font-size: 16px !important;           /* prevents iOS zoom on focus */
}

/* Metrics – stack nicely */
div[data-testid="stMetric"] {
    background: rgba(128,128,128,0.06);
    border-radius: 12px;
    padding: 12px 8px !important;
    text-align: center;
}

/* Expanders more readable */
.streamlit-expanderHeader {
    font-size: 1.05rem !important;
    font-weight: 600 !important;
}

/* Dataframes & tables */
.stDataFrame, .stTable {
    width: 100% !important;
    overflow-x: auto !important;
}

/* AgGrid container – allow horizontal scroll on small screens */
.ag-theme-streamlit, .ag-root-wrapper {
    width: 100% !important;
    max-width: 100% !important;
    overflow-x: auto !important;
    -webkit-overflow-scrolling: touch;
}

/* Images */
img {
    max-width: 100% !important;
    height: auto !important;
}

/* Captcha image specifically */
.stImage > img {
    max-width: min(100%, 420px) !important;
    max-height: 320px !important;
    width: auto !important;
    height: auto !important;
    object-fit: contain !important;
    margin: 0 auto;
    display: block;
}
/* Keep markdown header logo small even if CSS class is missed */
img.saral-logo,
.saral-header img {
    height: 48px !important;
    max-height: 48px !important;
    max-width: 120px !important;
    width: auto !important;
    object-fit: contain !important;
}

/* ---------- Header – compact, space-aware ---------- */
.saral-header {
    display: flex;
    align-items: center;
    padding: 8px 0 12px;
    margin-bottom: 12px;
    flex-wrap: wrap;
    gap: 10px;
}
.saral-logo {
    /* Strict size so the logo never dominates the screen */
    height: 48px !important;
    width: auto !important;
    max-height: 48px !important;
    max-width: 120px !important;
    border-radius: 8px;
    margin-right: 14px;
    object-fit: contain;
    flex-shrink: 0;
}
.saral-header-text { flex: 1; min-width: 160px; }
.saral-initiative { margin: 0; font-size: 1.0em; font-weight: 500; color: #4fc3f7; }
.saral-safety { color: #4fc3f7; font-weight: 700; }
.saral-title { margin: 2px 0 0; font-size: 1.75em; font-weight: bold; line-height: 1.15; }
.saral-subtitle { margin: 2px 0 0; font-size: 0.95em; color: #666; }

/* ---------- Footer credit ---------- */
.adaptive-credit {
    display: inline-block;
    padding: 12px 24px;
    background: var(--bg-glass, rgba(255,255,255,0.75));
    border: 2px solid #40c4ff;
    border-radius: 16px;
    backdrop-filter: blur(12px);
    -webkit-backdrop-filter: blur(12px);
    box-shadow: 0 6px 20px rgba(64,196,255,0.2);
    font-size: 14px;
    line-height: 1.45;
    max-width: 95vw;
}

/* ---------- Mobile / small tablet (<= 768px) ---------- */
@media (max-width: 768px) {
    .main .block-container {
        padding-left: 0.6rem !important;
        padding-right: 0.6rem !important;
        padding-top: 0.6rem !important;
    }

    /* Force columns to stack */
    div[data-testid="column"] {
        width: 100% !important;
        flex: 1 1 100% !important;
        min-width: 100% !important;
    }

    /* Header stacks */
    .saral-header {
        flex-direction: column;
        text-align: center;
        padding: 10px 0 16px;
    }
    .saral-logo {
        margin-right: 0;
        margin-bottom: 6px;
        height: 40px !important;
        max-height: 40px !important;
        max-width: 100px !important;
    }
    .saral-title { font-size: 1.45em; }
    .saral-subtitle { font-size: 0.85em; }
    .saral-initiative { font-size: 0.9em; }
    .saral-header {
        padding: 6px 0 10px;
        margin-bottom: 8px;
    }

    /* Metrics in a tighter grid */
    div[data-testid="stMetric"] {
        margin-bottom: 8px;
    }

    /* Buttons full-width on mobile for easier tapping */
    .stButton > button,
    .stDownloadButton > button {
        width: 100% !important;
    }

    /* Forms */
    div[data-testid="stForm"] {
        padding: 0.5rem !important;
    }

    /* Tabs – larger touch targets */
    button[data-baseweb="tab"] {
        font-size: 0.95rem !important;
        padding: 10px 12px !important;
    }

    /* Sidebar content a bit tighter */
    section[data-testid="stSidebar"] .block-container {
        padding-top: 1rem !important;
    }

    /* Reduce AgGrid height on mobile so user can still scroll the page */
    .ag-theme-streamlit {
        height: 420px !important;
    }

    /* WhatsApp / contact button */
    a button {
        font-size: 16px !important;
        padding: 12px 20px !important;
        width: 100% !important;
        max-width: 320px;
    }

    /* Captcha form columns */
    .stForm [data-testid="column"] {
        width: 100% !important;
    }
}

/* ---------- Very small phones (<= 400px) ---------- */
@media (max-width: 400px) {
    .saral-title { font-size: 1.65em; }
    .main .block-container {
        padding-left: 0.4rem !important;
        padding-right: 0.4rem !important;
    }
    .adaptive-credit {
        padding: 10px 14px;
        font-size: 13px;
    }
}

/* ---------- Tablet (769–1024px) – soft adjustments ---------- */
@media (min-width: 769px) and (max-width: 1024px) {
    .main .block-container {
        padding-left: 1.2rem !important;
        padding-right: 1.2rem !important;
    }
    .saral-title { font-size: 2.2em; }
}

/* ---------- Dark / light adaptive credit (kept from original) ---------- */
@media (prefers-color-scheme: light) {
  :root {
    --text-color: #1a1a1a; --text-highlight: #0d47a1; --text-sub: #1565c0;
    --bg-glass: rgba(255, 255, 255, 0.75); --border-color: #40c4ff;
    --shadow-base: rgba(64, 196, 255, 0.2); --shadow-hover: rgba(64, 196, 255, 0.35);
    --glow-color: rgba(179, 229, 252, 0.9);
  }
}
@media (prefers-color-scheme: dark) {
  :root {
    --text-color: #ffffff; --text-highlight: #e3f2fd; --text-sub: #bbdefb;
    --bg-glass: rgba(15, 25, 45, 0.65); --border-color: #40c4ff;
    --shadow-base: rgba(64, 196, 255, 0.15); --shadow-hover: rgba(64, 196, 255, 0.4);
    --glow-color: rgba(179, 229, 252, 0.95);
  }
}
.adaptive-credit p { margin: 0; color: var(--text-color); font-weight: 500; letter-spacing: 0.5px; }
.adaptive-credit p span.highlight { color: var(--text-highlight); font-weight: 700; }
.adaptive-credit p em { font-style: normal; color: var(--text-sub); }
.adaptive-credit:hover {
  transform: translateY(-3px);
  box-shadow: 0 12px 28px var(--shadow-hover), 0 0 30px var(--glow-color);
}
</style>
""",
    unsafe_allow_html=True,
)

# =========================================================================
# SESSION STATE INITIALIZATION (single pass — no duplicates)
# =========================================================================
DEFAULT_SESSION_STATE = {
    "captcha_text": lambda: None,       # the correct answer for the current image
    "captcha_fail_count": lambda: 0,    # wrong CAPTCHA attempts this session
    "logged_in": lambda: False,
    "user": lambda: None,
    "df": lambda: None,
    "feedback_submitting": lambda: False,
}
for _key, _default_factory in DEFAULT_SESSION_STATE.items():
    if _key not in st.session_state:
        st.session_state[_key] = _default_factory()


# =========================================================================
# IMAGE CAPTCHA
# =========================================================================
def _load_captcha_font(size):
    """Best-effort load of a bundled TrueType font; falls back to PIL's
    built-in bitmap font if none of the common system paths exist."""
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
        "C:\\Windows\\Fonts\\arialbd.ttf",
    ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            continue
    return ImageFont.load_default()


def generate_captcha_text(length=CAPTCHA_LENGTH):
    return "".join(random.choices(CAPTCHA_CHARS, k=length))


def generate_captcha_image(text, width=240, height=90):
    """Render `text` as a distorted image: randomised per-character
    position/rotation/colour, background noise lines and dots, plus a
    slight blur — enough to defeat simple OCR without being unreadable."""
    bg = (245, 247, 250)
    image = Image.new("RGB", (width, height), color=bg)
    draw = ImageDraw.Draw(image)

    # Background noise lines (drawn first, so text sits on top)
    for _ in range(8):
        xy = [
            (random.randint(0, width), random.randint(0, height)),
            (random.randint(0, width), random.randint(0, height)),
        ]
        draw.line(xy, fill=tuple(random.randint(190, 220) for _ in range(3)), width=2)

    font = _load_captcha_font(46)
    char_width = width // len(text)
    for i, ch in enumerate(text):
        char_img = Image.new("RGBA", (char_width, height), (0, 0, 0, 0))
        char_draw = ImageDraw.Draw(char_img)
        color = tuple(random.randint(20, 90) for _ in range(3))
        char_draw.text((char_width * 0.15, height * 0.2), ch, font=font, fill=color)
        angle = random.randint(-28, 28)
        rotated = char_img.rotate(angle, expand=0, resample=Image.BICUBIC)
        x_offset = i * char_width + random.randint(-4, 4)
        y_offset = random.randint(-6, 6)
        image.paste(rotated, (x_offset, y_offset), rotated)

    # Foreground noise dots
    for _ in range(120):
        x, y = random.randint(0, width - 1), random.randint(0, height - 1)
        draw.point((x, y), fill=tuple(random.randint(120, 180) for _ in range(3)))

    image = image.filter(ImageFilter.SMOOTH)
    buf = BytesIO()
    image.save(buf, format="PNG")
    buf.seek(0)
    return buf


def new_captcha():
    st.session_state.captcha_text = generate_captcha_text()


if st.session_state.captcha_text is None:
    new_captcha()


# =========================================================================
# LOGIN
# =========================================================================
def login(email, password):
    for user in st.secrets["users"]:
        if user["email"] == email and user["password"] == password:
            return user
    return None


if not st.session_state.logged_in:
    st.title("🔐 Login to S.A.R.A.L")
    st.caption("Safety Abnormality Report & Action List")

    if st.session_state.captcha_fail_count >= CAPTCHA_MAX_ATTEMPTS:
        st.error(
            "🚫 Too many incorrect attempts. Please refresh the page and try again."
        )
        st.stop()

    with st.form("login_form", clear_on_submit=False):
        email = st.text_input("📧 Email", placeholder="Enter Id")
        password = st.text_input("🔒 Password", type="password")

        st.markdown("**✅ Human check:**")
        st.image(
            generate_captcha_image(st.session_state.captcha_text),
            caption="Type the characters shown above (not case-sensitive)",
        )
        captcha_answer = st.text_input(
            "Enter the characters shown in the image",
            key="captcha_input",
            placeholder="e.g. AB3K9",
        )

        # On mobile these stack automatically thanks to CSS; on desktop they stay side-by-side
        btn_col1, btn_col2 = st.columns([2, 1])
        with btn_col1:
            submitted = st.form_submit_button("Login", use_container_width=True)
        with btn_col2:
            refresh_clicked = st.form_submit_button("🔄 New Image", use_container_width=True)

        if refresh_clicked:
            new_captcha()
            st.rerun()

        if submitted:
            given = (captcha_answer or "").strip().upper()
            expected = (st.session_state.captcha_text or "").upper()

            if not given or given != expected:
                st.session_state.captcha_fail_count += 1
                st.error("❌ Incorrect CAPTCHA. Please look at the new image and try again.")
                new_captcha()
                st.rerun()
            else:
                user = login(email, password)
                if user:
                    st.session_state.logged_in = True
                    st.session_state.user = user
                    st.session_state.captcha_fail_count = 0
                    new_captcha()  # burn this CAPTCHA so it can't be replayed
                    st.success(f"✅ Welcome, {user['name']}!")
                    st.rerun()
                else:
                    st.session_state.captcha_fail_count += 1
                    st.error("❌ Invalid email or password.")
                    new_captcha()
                    st.rerun()
    st.stop()

# ---------- ACKNOWLEDGMENT ----------
user_id = st.session_state.user["email"]  # use email as unique ID

try:
    ack_df = pd.read_excel("responses.xlsx")
    if "UserID" not in ack_df.columns or "Name" not in ack_df.columns:
        ack_df = pd.DataFrame(columns=["UserID", "Name"])
except FileNotFoundError:
    ack_df = pd.DataFrame(columns=["UserID", "Name"])

user_ack_done = user_id in ack_df["UserID"].values

if not user_ack_done:
    st.title("📢 Pending Deficiencies Compliance")
    with st.expander("⚠️ Pending Deficiencies Notice", expanded=True):
        st.info(
            """
            The compliance of deficiencies of previous dates are pending & needs to be completed immediately.
            I hereby declare that I have read this notice and will ensure compliance.
            """
        )
        with st.form("ack_form"):
            responder_name = st.text_input("✍️ Your Name")
            ack_submitted = st.form_submit_button("Submit Acknowledgment", use_container_width=True)
            if ack_submitted:
                if responder_name.strip():
                    new_entry = {"UserID": user_id, "Name": responder_name.strip()}
                    ack_df = pd.concat([ack_df, pd.DataFrame([new_entry])], ignore_index=True)
                    ack_df.to_excel("responses.xlsx", index=False)
                    st.success(f"✅ Thank you, {responder_name}, for acknowledging.")
                    st.rerun()
                else:
                    st.error("❌ Please enter your name before submitting.")
    st.stop()

# ---------- DISPLAY ALL RESPONSES ----------
st.markdown("### 📝 Responses Received")

try:
    ack_df = pd.read_excel("responses.xlsx")
    if not ack_df.empty:
        st.dataframe(ack_df, use_container_width=True, hide_index=True)
    else:
        st.info("No responses submitted yet.")
except FileNotFoundError:
    st.info("No responses submitted yet.")

# =========================================================================
# GOOGLE SHEETS CONNECTION
# =========================================================================
@st.cache_resource
def connect_to_gsheet():
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    service_account_info = dict(st.secrets["gcp_service_account"])
    if "private_key" in service_account_info:
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
    gc = gspread.authorize(creds)
    SHEET_ID = st.secrets["google_sheets"]["sheet_id"]
    SHEET_NAME = st.secrets["google_sheets"]["sheet_name"]
    return gc.open_by_key(SHEET_ID).worksheet(SHEET_NAME)


try:
    sheet = connect_to_gsheet()
    st.sidebar.success("✅ Connected to Google Sheets!")
except Exception as e:
    st.error(f"❌ Could not connect to Google Sheets: {e}")
    st.stop()


# =========================================================================
# INSPECTIONS SHEET CONNECTION (separate sheet, used for WhatsApp reports)
# =========================================================================
@st.cache_resource
def connect_to_inspections_sheet():
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    service_account_info = dict(st.secrets["gcp_service_account"])
    if "private_key" in service_account_info:
        service_account_info["private_key"] = service_account_info["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc.open_by_key(INSPECTIONS_SHEET_ID).worksheet(INSPECTIONS_SHEET_NAME)


@st.cache_data(ttl=30)
def load_inspections_data():
    """Load the 'Inspections' sheet using fixed column letters (see INSPECTIONS_COL_MAP)
    rather than header names, since the sheet header text may not match exactly."""
    cols = list(INSPECTIONS_COL_MAP.keys())
    try:
        ws = connect_to_inspections_sheet()
        values = ws.get_all_values()
        if not values or len(values) < 2:
            return pd.DataFrame(columns=cols + ["Date_parsed"])

        max_idx = max(INSPECTIONS_COL_MAP.values())
        records = []
        for row in values[1:]:  # skip header row
            if len(row) <= max_idx:
                row = row + [""] * (max_idx + 1 - len(row))
            records.append({name: row[idx].strip() for name, idx in INSPECTIONS_COL_MAP.items()})

        df = pd.DataFrame(records)
        # Drop fully blank rows
        df = df[df[cols].apply(lambda r: any(str(v).strip() for v in r), axis=1)].reset_index(drop=True)

        # Parse dates - try the default parser first, then fall back to day-first
        df["Date_parsed"] = pd.to_datetime(df["Date"], errors="coerce")
        mask = df["Date_parsed"].isna() & (df["Date"].astype(str).str.strip() != "")
        if mask.any():
            df.loc[mask, "Date_parsed"] = pd.to_datetime(df.loc[mask, "Date"], errors="coerce", dayfirst=True)

        return df
    except Exception as e:
        st.error(f"❌ Error loading Inspections sheet: {str(e)}")
        return pd.DataFrame(columns=cols + ["Date_parsed"])


def build_whatsapp_message(name, phone, designation_hq, insp_date_str, sections, submitted_dt=None):
    """Build a WhatsApp-formatted Safety Inspection Report matching the S.A.R.A.L template.

    sections: list of (type_of_inspection, location, deficiencies) tuples, where
              deficiencies is a list of (deficiency_text, action_by) tuples.
    """
    lines = [
        "*SAFETY INSPECTION REPORT*",
        "_Solapur Division • Central Railway_",
        f"*Inspecting Official(s):* {name or 'N/A'}",
        f"*Designation & HQ:* {designation_hq or 'N/A'}",
        f"*Date of Inspection:* {insp_date_str}",
    ]

    for i, (sec_type, sec_location, deficiencies) in enumerate(sections, start=1):
        lines.append(f"*{i}. {str(sec_type).strip().upper()}*")
        lines.append(str(sec_location).strip())
        lines.append("*Deficiencies:*")
        for j, (defect_text, action_by) in enumerate(deficiencies, start=1):
            lines.append(f"{j}. {defect_text}")
            lines.append(f"   Action By: {action_by if str(action_by).strip() else 'Not Assigned'}")

    lines.append("Photos: 0")
    lines.append(f"*Division:* Solapur • {phone or 'N/A'}")

    submitted_dt = submitted_dt or datetime.now(pytz.timezone("Asia/Kolkata"))
    time_part = submitted_dt.strftime("%I:%M:%S %p").lstrip("0").lower()
    submitted_str = f"{submitted_dt.day}/{submitted_dt.month}/{submitted_dt.year}, {time_part}"
    lines.append(f"*Submitted on:* {submitted_str}")
    lines.append(f"*Total Inspections:* {len(sections)}")
    lines.append("_Auto-generated via Safety Inspection App_")
    lines.append("Safety Department • Solapur Division")

    return "\n".join(lines)


# ---------- SIDEBAR ----------
st.sidebar.markdown(f"👤 Logged in as: **{st.session_state.user['name']}**")
st.sidebar.markdown(f"📧 {st.session_state.user['email']}")

if st.sidebar.button("🚪 Logout", use_container_width=True):
    st.session_state.logged_in = False
    st.session_state.user = None
    st.rerun()

# =========================================================================
# CONSTANT LISTS
# =========================================================================
STATION_LIST = list(dict.fromkeys([
    'BRB', 'MLM', 'BGVN', 'JNTR', 'KEU', 'WSB', 'PPJ', 'JEUR', 'KEM', 'BLNI', 'DHS', 'KWV', 'WDS', 'MA', 'AAG',
    'MKPT', 'MO', 'MVE', 'PK', 'BALE', "SUR", 'TKWD', 'HG', 'TLT', 'AKOR', 'NGS', 'BOT', 'DUD', 'KUI', 'GDGN', 'GUR',
    'HHD', 'SVG', 'BBD', 'TJSP', 'KLBG', 'HQR', 'MR', 'SDB', 'WADI', 'ARAG', 'BLNK', 'SGRE', 'KVK', 'LNP', 'DLGN',
    'JTRD', 'MSDG', 'JVA', 'WSD', 'SGLA', 'PVR', 'MLB', 'SEI', 'BTW', 'PJR', 'DRSV', 'YSI', 'KMRD', 'DKY', 'MRX',
    'OSA', 'HGL', 'LUR', 'NTPC', 'MRJ', 'BHLI'
]))

GATE_LIST = list(dict.fromkeys([
    'LC-19', 'LC-22A', 'LC-25', 'LC-26', 'LC-27C', 'LC-28', 'LC-30', 'LC-31', 'LC-35', 'LC-37', 'LC-40', 'LC-41',
    'LC-43', 'LC-44', 'LC-45', 'LC-46C', 'LC-54', 'LC-61', 'LC-66', 'LC-74', 'LC-76', 'LC-78', 'LC-82', 'LC-1',
    'LC-60A', 'LC-1 ACC', 'LC-2 ACC', 'LC-91', 'LC-22', 'LC-24', 'LC-32', 'LC-49', 'LC-70',
    'LC-10', 'LC-34', 'LC-36', 'LC-47', 'LC-55', 'LC-59', 'LC-2', 'LC-4', 'LC-42', 'LC-02', 'LC-128', 'LC-63',
    'LC-04', 'LC-67', 'LC-77', 'LC-75', 'LC-64', 'LC-65', 'LC-5', 'LC-6', 'LC-57', 'LC-62', 'LC-39', 'LC-2/C',
    'LC-6/C', 'LC-11', 'LC-03', 'LC-15/C', 'LC-21', 'LC-26-A', 'LC-60'
]))

FOOTPLATE_ROUTE_HIERARCHY = {
    "SUR-DD": [
        "SUR-BALE", "SUR-PK", "SUR-MVE", "SUR-MO", "SUR-MKPT", "SUR-WKA", "SUR-ANG", "SUR-MA",
        "SUR-WDS", "SUR-KWV", "SUR-KEM", "SUR-DHS", "SUR-BLNI", "SUR-JEUR", "SUR-PPJ", "SUR-WSB",
        "SUR-KEU", "SUR-JNTR", "SUR-BGVN", "SUR-MLM", "SUR-BRB", "SUR-DD", "BALE-PK", "BALE-MVE",
        "BALE-MO", "BALE-MKPT", "BALE-WKA", "BALE-ANG", "BALE-MA", "BALE-WDS", "BALE-KWV", "BALE-KEM",
        "BALE-DHS", "BALE-BLNI", "BALE-JEUR", "BALE-PPJ", "BALE-WSB", "BALE-KEU", "BALE-JNTR", "BALE-BGVN",
        "BALE-MLM", "BALE-BRB", "BALE-DD", "PK-MVE", "PK-MO", "PK-MKPT", "PK-WKA", "PK-ANG",
        "PK-MA", "PK-WDS", "PK-KWV", "PK-KEM", "PK-DHS", "PK-BLNI", "PK-JEUR", "PK-PPJ",
        "PK-WSB", "PK-KEU", "PK-JNTR", "PK-BGVN", "PK-MLM", "PK-BRB", "PK-DD", "MVE-MO",
        "MVE-MKPT", "MVE-WKA", "MVE-ANG", "MVE-MA", "MVE-WDS", "MVE-KWV", "MVE-KEM", "MVE-DHS",
        "MVE-BLNI", "MVE-JEUR", "MVE-PPJ", "MVE-WSB", "MVE-KEU", "MVE-JNTR", "MVE-BGVN", "MVE-MLM",
        "MVE-BRB", "MVE-DD", "MO-MKPT", "MO-WKA", "MO-ANG", "MO-MA", "MO-WDS", "MO-KWV",
        "MO-KEM", "MO-DHS", "MO-BLNI", "MO-JEUR", "MO-PPJ", "MO-WSB", "MO-KEU", "MO-JNTR",
        "MO-BGVN", "MO-MLM", "MO-BRB", "MO-DD", "MKPT-WKA", "MKPT-ANG", "MKPT-MA", "MKPT-WDS",
        "MKPT-KWV", "MKPT-KEM", "MKPT-DHS", "MKPT-BLNI", "MKPT-JEUR", "MKPT-PPJ", "MKPT-WSB", "MKPT-KEU",
        "MKPT-JNTR", "MKPT-BGVN", "MKPT-MLM", "MKPT-BRB", "MKPT-DD", "WKA-ANG", "WKA-MA", "WKA-WDS",
        "WKA-KWV", "WKA-KEM", "WKA-DHS", "WKA-BLNI", "WKA-JEUR", "WKA-PPJ", "WKA-WSB", "WKA-KEU",
        "WKA-JNTR", "WKA-BGVN", "WKA-MLM", "WKA-BRB", "WKA-DD", "ANG-MA", "ANG-WDS", "ANG-KWV",
        "ANG-KEM", "ANG-DHS", "ANG-BLNI", "ANG-JEUR", "ANG-PPJ", "ANG-WSB", "ANG-KEU", "ANG-JNTR",
        "ANG-BGVN", "ANG-MLM", "ANG-BRB", "ANG-DD", "MA-WDS", "MA-KWV", "MA-KEM", "MA-DHS",
        "MA-BLNI", "MA-JEUR", "MA-PPJ", "MA-WSB", "MA-KEU", "MA-JNTR", "MA-BGVN", "MA-MLM",
        "MA-BRB", "MA-DD", "WDS-KWV", "WDS-KEM", "WDS-DHS", "WDS-BLNI", "WDS-JEUR", "WDS-PPJ",
        "WDS-WSB", "WDS-KEU", "WDS-JNTR", "WDS-BGVN", "WDS-MLM", "WDS-BRB", "WDS-DD", "KWV-KEM",
        "KWV-DHS", "KWV-BLNI", "KWV-JEUR", "KWV-PPJ", "KWV-WSB", "KWV-KEU", "KWV-JNTR", "KWV-BGVN",
        "KWV-MLM", "KWV-BRB", "KWV-DD", "KEM-DHS", "KEM-BLNI", "KEM-JEUR", "KEM-PPJ", "KEM-WSB",
        "KEM-KEU", "KEM-JNTR", "KEM-BGVN", "KEM-MLM", "KEM-BRB", "KEM-DD", "DHS-BLNI", "DHS-JEUR",
        "DHS-PPJ", "DHS-WSB", "DHS-KEU", "DHS-JNTR", "DHS-BGVN", "DHS-MLM", "DHS-BRB", "DHS-DD",
        "BLNI-JEUR", "BLNI-PPJ", "BLNI-WSB", "BLNI-KEU", "BLNI-JNTR", "BLNI-BGVN", "BLNI-MLM", "BLNI-BRB",
        "BLNI-DD", "JEUR-PPJ", "JEUR-WSB", "JEUR-KEU", "JEUR-JNTR", "JEUR-BGVN", "JEUR-MLM", "JEUR-BRB",
        "JEUR-DD", "PPJ-WSB", "PPJ-KEU", "PPJ-JNTR", "PPJ-BGVN", "PPJ-MLM", "PPJ-BRB", "PPJ-DD",
        "WSB-KEU", "WSB-JNTR", "WSB-BGVN", "WSB-MLM", "WSB-BRB", "WSB-DD", "KEU-JNTR", "KEU-BGVN",
        "KEU-MLM", "KEU-BRB", "KEU-DD", "JNTR-BGVN", "JNTR-MLM", "JNTR-BRB", "JNTR-DD", "BGVN-MLM",
        "BGVN-BRB", "BGVN-DD", "MLM-BRB", "MLM-DD", "BRB-DD", 'SUR', 'BALE', 'PK', 'MVE', 'MO', 'MKPT', 'WKA', 'ANG', 'MA', 'WDS', 'KWV', 'KEM', 'DHS', 'BLNI', 'JEUR', 'PPJ', 'WSB', 'KEU', 'JNTR', 'BGVN', 'MLM', 'BRB', 'DD', 'LC-40', 'LC-42', 'LC-21', 'LC-19'
    ],
    "DD-SUR": [
        "DD-BRB", "DD-MLM", "DD-BGVN", "DD-JNTR", "DD-KEU", "DD-WSB", "DD-PPJ", "DD-JEUR",
        "DD-BLNI", "DD-DHS", "DD-KEM", "DD-KWV", "DD-WDS", "DD-MA", "DD-ANG", "DD-WKA",
        "DD-MKPT", "DD-MO", "DD-MVE", "DD-PK", "DD-BALE", "DD-SUR", "BRB-MLM", "BRB-BGVN",
        "BRB-JNTR", "BRB-KEU", "BRB-WSB", "BRB-PPJ", "BRB-JEUR", "BRB-BLNI", "BRB-DHS", "BRB-KEM",
        "BRB-KWV", "BRB-WDS", "BRB-MA", "BRB-ANG", "BRB-WKA", "BRB-MKPT", "BRB-MO", "BRB-MVE",
        "BRB-PK", "BRB-BALE", "BRB-SUR", "MLM-BGVN", "MLM-JNTR", "MLM-KEU", "MLM-WSB", "MLM-PPJ",
        "MLM-JEUR", "MLM-BLNI", "MLM-DHS", "MLM-KEM", "MLM-KWV", "MLM-WDS", "MLM-MA", "MLM-ANG",
        "MLM-WKA", "MLM-MKPT", "MLM-MO", "MLM-MVE", "MLM-PK", "MLM-BALE", "MLM-SUR", "BGVN-JNTR",
        "BGVN-KEU", "BGVN-WSB", "BGVN-PPJ", "BGVN-JEUR", "BGVN-BLNI", "BGVN-DHS", "BGVN-KEM", "BGVN-KWV",
        "BGVN-WDS", "BGVN-MA", "BGVN-ANG", "BGVN-WKA", "BGVN-MKPT", "BGVN-MO", "BGVN-MVE", "BGVN-PK",
        "BGVN-BALE", "BGVN-SUR", "JNTR-KEU", "JNTR-WSB", "JNTR-PPJ", "JNTR-JEUR", "JNTR-BLNI", "JNTR-DHS",
        "JNTR-KEM", "JNTR-KWV", "JNTR-WDS", "JNTR-MA", "JNTR-ANG", "JNTR-WKA", "JNTR-MKPT", "JNTR-MO",
        "JNTR-MVE", "JNTR-PK", "JNTR-BALE", "JNTR-SUR", "KEU-WSB", "KEU-PPJ", "KEU-JEUR", "KEU-BLNI",
        "KEU-DHS", "KEU-KEM", "KEU-KWV", "KEU-WDS", "KEU-MA", "KEU-ANG", "KEU-WKA", "KEU-MKPT",
        "KEU-MO", "KEU-MVE", "KEU-PK", "KEU-BALE", "KEU-SUR", "WSB-PPJ", "WSB-JEUR", "WSB-BLNI",
        "WSB-DHS", "WSB-KEM", "WSB-KWV", "WSB-WDS", "WSB-MA", "WSB-ANG", "WSB-WKA", "WSB-MKPT",
        "WSB-MO", "WSB-MVE", "WSB-PK", "WSB-BALE", "WSB-SUR", "PPJ-JEUR", "PPJ-BLNI", "PPJ-DHS",
        "PPJ-KEM", "PPJ-KWV", "PPJ-WDS", "PPJ-MA", "PPJ-ANG", "PPJ-WKA", "PPJ-MKPT", "PPJ-MO",
        "PPJ-MVE", "PPJ-PK", "PPJ-BALE", "PPJ-SUR", "JEUR-BLNI", "JEUR-DHS", "JEUR-KEM", "JEUR-KWV",
        "JEUR-WDS", "JEUR-MA", "JEUR-ANG", "JEUR-WKA", "JEUR-MKPT", "JEUR-MO", "JEUR-MVE", "JEUR-PK",
        "JEUR-BALE", "JEUR-SUR", "BLNI-DHS", "BLNI-KEM", "BLNI-KWV", "BLNI-WDS", "BLNI-MA", "BLNI-ANG",
        "BLNI-WKA", "BLNI-MKPT", "BLNI-MO", "BLNI-MVE", "BLNI-PK", "BLNI-BALE", "BLNI-SUR", "DHS-KEM",
        "DHS-KWV", "DHS-WDS", "DHS-MA", "DHS-ANG", "DHS-WKA", "DHS-MKPT", "DHS-MO", "DHS-MVE",
        "DHS-PK", "DHS-BALE", "DHS-SUR", "KEM-KWV", "KEM-WDS", "KEM-MA", "KEM-ANG", "KEM-WKA",
        "KEM-MKPT", "KEM-MO", "KEM-MVE", "KEM-PK", "KEM-BALE", "KEM-SUR", "KWV-WDS", "KWV-MA",
        "KWV-ANG", "KWV-WKA", "KWV-MKPT", "KWV-MO", "KWV-MVE", "KWV-PK", "KWV-BALE", "KWV-SUR",
        "WDS-MA", "WDS-ANG", "WDS-WKA", "WDS-MKPT", "WDS-MO", "WDS-MVE", "WDS-PK", "WDS-BALE",
        "WDS-SUR", "MA-ANG", "MA-WKA", "MA-MKPT", "MA-MO", "MA-MVE", "MA-PK", "MA-BALE",
        "MA-SUR", "ANG-WKA", "ANG-MKPT", "ANG-MO", "ANG-MVE", "ANG-PK", "ANG-BALE", "ANG-SUR",
        "WKA-MKPT", "WKA-MO", "WKA-MVE", "WKA-PK", "WKA-BALE", "WKA-SUR", "MKPT-MO", "MKPT-MVE",
        "MKPT-PK", "MKPT-BALE", "MKPT-SUR", "MO-MVE", "MO-PK", "MO-BALE", "MO-SUR", "MVE-PK",
        "MVE-BALE", "MVE-SUR", "PK-BALE", "PK-SUR", "BALE-SUR", 'SUR', 'BALE', 'PK', 'MVE', 'MO', 'MKPT', 'WKA', 'ANG', 'MA', 'WDS', 'KWV', 'KEM', 'DHS', 'BLNI', 'JEUR', 'PPJ', 'WSB', 'KEU', 'JNTR', 'BGVN', 'MLM', 'BRB', 'DD', 'LC-40', 'LC-42', 'LC-21', 'LC-19'
    ],
    "SUR-WADI": [
        "SUR-TKWD", "SUR-HG", "SUR-TLT", "SUR-AKOR", "SUR-NGS", "SUR-BOT", "SUR-GUR", "SUR-GDGN",
        "SUR-KUI", "SUR-DUD", "SUR-HDD", "SUR-SVG", "SUR-BBD", "SUR-TJSP", "SUR-KLBG", "SUR-HQR",
        "SUR-MR", "SUR-SDB", "SUR-WADI", "TKWD-HG", "TKWD-TLT", "TKWD-AKOR", "TKWD-NGS", "TKWD-BOT",
        "TKWD-GUR", "TKWD-GDGN", "TKWD-KUI", "TKWD-DUD", "TKWD-HDD", "TKWD-SVG", "TKWD-BBD", "TKWD-TJSP",
        "TKWD-KLBG", "TKWD-HQR", "TKWD-MR", "TKWD-SDB", "TKWD-WADI", "HG-TLT", "HG-AKOR", "HG-NGS",
        "HG-BOT", "HG-GUR", "HG-GDGN", "HG-KUI", "HG-DUD", "HG-HDD", "HG-SVG", "HG-BBD",
        "HG-TJSP", "HG-KLBG", "HG-HQR", "HG-MR", "HG-SDB", "HG-WADI", "TLT-AKOR", "TLT-NGS",
        "TLT-BOT", "TLT-GUR", "TLT-GDGN", "TLT-KUI", "TLT-DUD", "TLT-HDD", "TLT-SVG", "TLT-BBD",
        "TLT-TJSP", "TLT-KLBG", "TLT-HQR", "TLT-MR", "TLT-SDB", "TLT-WADI", "AKOR-NGS", "AKOR-BOT",
        "AKOR-GUR", "AKOR-GDGN", "AKOR-KUI", "AKOR-DUD", "AKOR-HDD", "AKOR-SVG", "AKOR-BBD", "AKOR-TJSP",
        "AKOR-KLBG", "AKOR-HQR", "AKOR-MR", "AKOR-SDB", "AKOR-WADI", "NGS-BOT", "NGS-GUR", "NGS-GDGN",
        "NGS-KUI", "NGS-DUD", "NGS-HDD", "NGS-SVG", "NGS-BBD", "NGS-TJSP", "NGS-KLBG", "NGS-HQR",
        "NGS-MR", "NGS-SDB", "NGS-WADI", "BOT-GUR", "BOT-GDGN", "BOT-KUI", "BOT-DUD", "BOT-HDD",
        "BOT-SVG", "BOT-BBD", "BOT-TJSP", "BOT-KLBG", "BOT-HQR", "BOT-MR", "BOT-SDB", "BOT-WADI",
        "GUR-GDGN", "GUR-KUI", "GUR-DUD", "GUR-HDD", "GUR-SVG", "GUR-BBD", "GUR-TJSP", "GUR-KLBG",
        "GUR-HQR", "GUR-MR", "GUR-SDB", "GUR-WADI", "GDGN-KUI", "GDGN-DUD", "GDGN-HDD", "GDGN-SVG",
        "GDGN-BBD", "GDGN-TJSP", "GDGN-KLBG", "GDGN-HQR", "GDGN-MR", "GDGN-SDB", "GDGN-WADI", "KUI-DUD",
        "KUI-HDD", "KUI-SVG", "KUI-BBD", "KUI-TJSP", "KUI-KLBG", "KUI-HQR", "KUI-MR", "KUI-SDB",
        "KUI-WADI", "DUD-HDD", "DUD-SVG", "DUD-BBD", "DUD-TJSP", "DUD-KLBG", "DUD-HQR", "DUD-MR",
        "DUD-SDB", "DUD-WADI", "HDD-SVG", "HDD-BBD", "HDD-TJSP", "HDD-KLBG", "HDD-HQR", "HDD-MR",
        "HDD-SDB", "HDD-WADI", "SVG-BBD", "SVG-TJSP", "SVG-KLBG", "SVG-HQR", "SVG-MR", "SVG-SDB",
        "SVG-WADI", "BBD-TJSP", "BBD-KLBG", "BBD-HQR", "BBD-MR", "BBD-SDB", "BBD-WADI", "TJSP-KLBG",
        "TJSP-HQR", "TJSP-MR", "TJSP-SDB", "TJSP-WADI", "KLBG-HQR", "KLBG-MR", "KLBG-SDB", "KLBG-WADI",
        "HQR-MR", "HQR-SDB", "HQR-WADI", "MR-SDB", "MR-WADI", "SDB-WADI", 'SUR', 'TKWD', 'HG', 'TLT', 'AKOR', 'NGS', 'BOT', 'GUR', 'GDGN', 
        'KUI', 'DUD', 'HDD', 'SVG', 'BBD', 'TJSP', 'KLBG', 'HQR', 'MR', 'SDB', 'WADI', 'LC-1', 'LC-60', 'LC-61', 'LC-66', 'LC-74', 'LC-82', 'LC-91'
    ],
    "WADI-SUR": [
        "WADI-SDB", "WADI-MR", "WADI-HQR", "WADI-KLBG", "WADI-TJSP", "WADI-BBD", "WADI-SVG", "WADI-HDD",
        "WADI-DUD", "WADI-KUI", "WADI-GDGN", "WADI-GUR", "WADI-BOT", "WADI-NGS", "WADI-AKOR", "WADI-TLT",
        "WADI-HG", "WADI-TKWD", "WADI-SUR", "SDB-MR", "SDB-HQR", "SDB-KLBG", "SDB-TJSP", "SDB-BBD",
        "SDB-SVG", "SDB-HDD", "SDB-DUD", "SDB-KUI", "SDB-GDGN", "SDB-GUR", "SDB-BOT", "SDB-NGS",
        "SDB-AKOR", "SDB-TLT", "SDB-HG", "SDB-TKWD", "SDB-SUR", "MR-HQR", "MR-KLBG", "MR-TJSP",
        "MR-BBD", "MR-SVG", "MR-HDD", "MR-DUD", "MR-KUI", "MR-GDGN", "MR-GUR", "MR-BOT",
        "MR-NGS", "MR-AKOR", "MR-TLT", "MR-HG", "MR-TKWD", "MR-SUR", "HQR-KLBG", "HQR-TJSP",
        "HQR-BBD", "HQR-SVG", "HQR-HDD", "HQR-DUD", "HQR-KUI", "HQR-GDGN", "HQR-GUR", "HQR-BOT",
        "HQR-NGS", "HQR-AKOR", "HQR-TLT", "HQR-HG", "HQR-TKWD", "HQR-SUR", "KLBG-TJSP", "KLBG-BBD",
        "KLBG-SVG", "KLBG-HDD", "KLBG-DUD", "KLBG-KUI", "KLBG-GDGN", "KLBG-GUR", "KLBG-BOT", "KLBG-NGS",
        "KLBG-AKOR", "KLBG-TLT", "KLBG-HG", "KLBG-TKWD", "KLBG-SUR", "TJSP-BBD", "TJSP-SVG", "TJSP-HDD",
        "TJSP-DUD", "TJSP-KUI", "TJSP-GDGN", "TJSP-GUR", "TJSP-BOT", "TJSP-NGS", "TJSP-AKOR", "TJSP-TLT",
        "TJSP-HG", "TJSP-TKWD", "TJSP-SUR", "BBD-SVG", "BBD-HDD", "BBD-DUD", "BBD-KUI", "BBD-GDGN",
        "BBD-GUR", "BBD-BOT", "BBD-NGS", "BBD-AKOR", "BBD-TLT", "BBD-HG", "BBD-TKWD", "BBD-SUR",
        "SVG-HDD", "SVG-DUD", "SVG-KUI", "SVG-GDGN", "SVG-GUR", "SVG-BOT", "SVG-NGS", "SVG-AKOR",
        "SVG-TLT", "SVG-HG", "SVG-TKWD", "SVG-SUR", "HDD-DUD", "HDD-KUI", "HDD-GDGN", "HDD-GUR",
        "HDD-BOT", "HDD-NGS", "HDD-AKOR", "HDD-TLT", "HDD-HG", "HDD-TKWD", "HDD-SUR", "DUD-KUI",
        "DUD-GDGN", "DUD-GUR", "DUD-BOT", "DUD-NGS", "DUD-AKOR", "DUD-TLT", "DUD-HG", "DUD-TKWD",
        "DUD-SUR", "KUI-GDGN", "KUI-GUR", "KUI-BOT", "KUI-NGS", "KUI-AKOR", "KUI-TLT", "KUI-HG",
        "KUI-TKWD", "KUI-SUR", "GDGN-GUR", "GDGN-BOT", "GDGN-NGS", "GDGN-AKOR", "GDGN-TLT", "GDGN-HG",
        "GDGN-TKWD", "GDGN-SUR", "GUR-BOT", "GUR-NGS", "GUR-AKOR", "GUR-TLT", "GUR-HG", "GUR-TKWD",
        "GUR-SUR", "BOT-NGS", "BOT-AKOR", "BOT-TLT", "BOT-HG", "BOT-TKWD", "BOT-SUR", "NGS-AKOR",
        "NGS-TLT", "NGS-HG", "NGS-TKWD", "NGS-SUR", "AKOR-TLT", "AKOR-HG", "AKOR-TKWD", "AKOR-SUR",
        "TLT-HG", "TLT-TKWD", "TLT-SUR", "HG-TKWD", "HG-SUR", "TKWD-SUR", 'SUR', 'TKWD', 'HG', 'TLT', 'AKOR', 'NGS', 'BOT', 'GUR', 'GDGN', 'KUI', 'DUD', 'HDD', 'SVG', 'BBD', 'TJSP', 'KLBG', 'HQR', 'MR', 'SDB', 'WADI', 'LC-1', 'LC-60', 
        'LC-61', 'LC-66', 'LC-74', 'LC-82', 'LC-91'
    ],
    "LUR-KWV": [
        "LUR-HGL", "LUR-OSA", "LUR-MRX", "LUR-DKY", "LUR-KMRD", "LUR-YSI", "LUR-DRSV", "LUR-PJR",
        "LUR-BTW", "LUR-SEI", "LUR-KWV", "HGL-OSA", "HGL-MRX", "HGL-DKY", "HGL-KMRD", "HGL-YSI",
        "HGL-DRSV", "HGL-PJR", "HGL-BTW", "HGL-SEI", "HGL-KWV", "OSA-MRX", "OSA-DKY", "OSA-KMRD",
        "OSA-YSI", "OSA-DRSV", "OSA-PJR", "OSA-BTW", "OSA-SEI", "OSA-KWV", "MRX-DKY", "MRX-KMRD",
        "MRX-YSI", "MRX-DRSV", "MRX-PJR", "MRX-BTW", "MRX-SEI", "MRX-KWV", "DKY-KMRD", "DKY-YSI",
        "DKY-DRSV", "DKY-PJR", "DKY-BTW", "DKY-SEI", "DKY-KWV", "KMRD-YSI", "KMRD-DRSV", "KMRD-PJR",
        "KMRD-BTW", "KMRD-SEI", "KMRD-KWV", "YSI-DRSV", "YSI-PJR", "YSI-BTW", "YSI-SEI", "YSI-KWV",
        "DRSV-PJR", "DRSV-BTW", "DRSV-SEI", "DRSV-KWV", "PJR-BTW", "PJR-SEI", "PJR-KWV", "BTW-SEI",
        "BTW-KWV", "SEI-KWV", 'SEI', 'BTW', 'PJR', 'DRSV', 'YSI', 'KMRD', 'DKY', 'MRX', 'OSA', 'HGL', 'LUR'
    ],
    "KWV-LUR": [
        "KWV-SEI", "KWV-BTW", "KWV-PJR", "KWV-DRSV", "KWV-YSI", "KWV-KMRD", "KWV-DKY", "KWV-MRX",
        "KWV-OSA", "KWV-HGL", "KWV-LUR", "SEI-BTW", "SEI-PJR", "SEI-DRSV", "SEI-YSI", "SEI-KMRD",
        "SEI-DKY", "SEI-MRX", "SEI-OSA", "SEI-HGL", "SEI-LUR", "BTW-PJR", "BTW-DRSV", "BTW-YSI",
        "BTW-KMRD", "BTW-DKY", "BTW-MRX", "BTW-OSA", "BTW-HGL", "BTW-LUR", "PJR-DRSV", "PJR-YSI",
        "PJR-KMRD", "PJR-DKY", "PJR-MRX", "PJR-OSA", "PJR-HGL", "PJR-LUR", "DRSV-YSI", "DRSV-KMRD",
        "DRSV-DKY", "DRSV-MRX", "DRSV-OSA", "DRSV-HGL", "DRSV-LUR", "YSI-KMRD", "YSI-DKY", "YSI-MRX",
        "YSI-OSA", "YSI-HGL", "YSI-LUR", "KMRD-DKY", "KMRD-MRX", "KMRD-OSA", "KMRD-HGL", "KMRD-LUR",
        "DKY-MRX", "DKY-OSA", "DKY-HGL", "DKY-LUR", "MRX-OSA", "MRX-HGL", "MRX-LUR", "OSA-HGL",
        "OSA-LUR", "HGL-LUR", 'SEI', 'BTW', 'PJR', 'DRSV', 'YSI', 'KMRD', 'DKY', 'MRX', 'OSA', 'HGL', 'LUR'
    ],
    "KWV-MRJ": [
        "KWV-ARAG", "KWV-BLNK", "KWV-SGRE", "KWV-KVK", "KWV-LNP", "KWV-DLGN", "KWV-JTRD", "KWV-MSDG",
        "KWV-JVA", "KWV-WSD", "KWV-SGLA", "KWV-PVR", "KWV-MLB", "KWV-MRJ", "ARAG-BLNK", "ARAG-SGRE",
        "ARAG-KVK", "ARAG-LNP", "ARAG-DLGN", "ARAG-JTRD", "ARAG-MSDG", "ARAG-JVA", "ARAG-WSD", "ARAG-SGLA",
        "ARAG-PVR", "ARAG-MLB", "ARAG-MRJ", "BLNK-SGRE", "BLNK-KVK", "BLNK-LNP", "BLNK-DLGN", "BLNK-JTRD",
        "BLNK-MSDG", "BLNK-JVA", "BLNK-WSD", "BLNK-SGLA", "BLNK-PVR", "BLNK-MLB", "BLNK-MRJ", "SGRE-KVK",
        "SGRE-LNP", "SGRE-DLGN", "SGRE-JTRD", "SGRE-MSDG", "SGRE-JVA", "SGRE-WSD", "SGRE-SGLA", "SGRE-PVR",
        "SGRE-MLB", "SGRE-MRJ", "KVK-LNP", "KVK-DLGN", "KVK-JTRD", "KVK-MSDG", "KVK-JVA", "KVK-WSD",
        "KVK-SGLA", "KVK-PVR", "KVK-MLB", "KVK-MRJ", "LNP-DLGN", "LNP-JTRD", "LNP-MSDG", "LNP-JVA",
        "LNP-WSD", "LNP-SGLA", "LNP-PVR", "LNP-MLB", "LNP-MRJ", "DLGN-JTRD", "DLGN-MSDG", "DLGN-JVA",
        "DLGN-WSD", "DLGN-SGLA", "DLGN-PVR", "DLGN-MLB", "DLGN-MRJ", "JTRD-MSDG", "JTRD-JVA", "JTRD-WSD",
        "JTRD-SGLA", "JTRD-PVR", "JTRD-MLB", "JTRD-MRJ", "MSDG-JVA", "MSDG-WSD", "MSDG-SGLA", "MSDG-PVR",
        "MSDG-MLB", "MSDG-MRJ", "JVA-WSD", "JVA-SGLA", "JVA-PVR", "JVA-MLB", "JVA-MRJ", "WSD-SGLA",
        "WSD-PVR", "WSD-MLB", "WSD-MRJ", "SGLA-PVR", "SGLA-MLB", "SGLA-MRJ", "PVR-MLB", "PVR-MRJ",
        "MLB-MRJ", 'ARAG', 'BLNK', 'SGRE', 'KVK', 'LNP', 'DLGN', 'JTRD', 'MSDG', 'JVA', 'WSD', 'SGLA', 'PVR', 'MLB'
    ],
    "MRJ-KWV": [
        "MRJ-MLB", "MRJ-PVR", "MRJ-SGLA", "MRJ-WSD", "MRJ-JVA", "MRJ-MSDG", "MRJ-JTRD", "MRJ-DLGN",
        "MRJ-LNP", "MRJ-KVK", "MRJ-SGRE", "MRJ-BLNK", "MRJ-ARAG", "MRJ-KWV", "MLB-PVR", "MLB-SGLA",
        "MLB-WSD", "MLB-JVA", "MLB-MSDG", "MLB-JTRD", "MLB-DLGN", "MLB-LNP", "MLB-KVK", "MLB-SGRE",
        "MLB-BLNK", "MLB-ARAG", "MLB-KWV", "PVR-SGLA", "PVR-WSD", "PVR-JVA", "PVR-MSDG", "PVR-JTRD",
        "PVR-DLGN", "PVR-LNP", "PVR-KVK", "PVR-SGRE", "PVR-BLNK", "PVR-ARAG", "PVR-KWV", "SGLA-WSD",
        "SGLA-JVA", "SGLA-MSDG", "SGLA-JTRD", "SGLA-DLGN", "SGLA-LNP", "SGLA-KVK", "SGLA-SGRE", "SGLA-BLNK",
        "SGLA-ARAG", "SGLA-KWV", "WSD-JVA", "WSD-MSDG", "WSD-JTRD", "WSD-DLGN", "WSD-LNP", "WSD-KVK",
        "WSD-SGRE", "WSD-BLNK", "WSD-ARAG", "WSD-KWV", "JVA-MSDG", "JVA-JTRD", "JVA-DLGN", "JVA-LNP",
        "JVA-KVK", "JVA-SGRE", "JVA-BLNK", "JVA-ARAG", "JVA-KWV", "MSDG-JTRD", "MSDG-DLGN", "MSDG-LNP",
        "MSDG-KVK", "MSDG-SGRE", "MSDG-BLNK", "MSDG-ARAG", "MSDG-KWV", "JTRD-DLGN", "JTRD-LNP", "JTRD-KVK",
        "JTRD-SGRE", "JTRD-BLNK", "JTRD-ARAG", "JTRD-KWV", "DLGN-LNP", "DLGN-KVK", "DLGN-SGRE", "DLGN-BLNK",
        "DLGN-ARAG", "DLGN-KWV", "LNP-KVK", "LNP-SGRE", "LNP-BLNK", "LNP-ARAG", "LNP-KWV", "KVK-SGRE",
        "KVK-BLNK", "KVK-ARAG", "KVK-KWV", "SGRE-BLNK", "SGRE-ARAG", "SGRE-KWV", "BLNK-ARAG", "BLNK-KWV",
        "ARAG-KWV", 'ARAG', 'BLNK', 'SGRE', 'KVK', 'LNP', 'DLGN', 'JTRD', 'MSDG', 'JVA', 'WSD', 'SGLA', 'PVR', 'MLB'
    ]
}

FOOTPLATE_ROUTES = list(FOOTPLATE_ROUTE_HIERARCHY.keys())
ALL_FOOTPLATE_LOCATIONS = FOOTPLATE_ROUTES + [sub for subs in FOOTPLATE_ROUTE_HIERARCHY.values() for sub in subs]
ALL_LOCATIONS = STATION_LIST + GATE_LIST + ALL_FOOTPLATE_LOCATIONS

HEAD_LIST = ["", "ELECT/TRD", "ELECT/G", "ELECT/TRO", "SIGNAL & TELECOM", "OPTG", "MECHANICAL",
             "ENGINEERING", "COMMERCIAL", 'PERSONNEL', 'SECURITY', "FINANCE", "MEDICAL", "STORE", 'GSU']

SUBHEAD_LIST = {
    "ELECT/TRD": ["T/W WAGON", "TSS/SP/SSP", "OHE SECTION", "OHE STATION", "MISC"],
    "ELECT/G": ["TL/AC COACH", "POWER/PANTRY CAR", "WIRING/EQUIPMENT", "UPS", "AC", "DG", "SOLAR LIGHT", "MISC", 'LIGHT/ILLUMINATION'],
    "ELECT/TRO": ["LOCO DEFECTS", "RUNNING ROOM DEFICIENCIES", "LOBBY DEFICIENCIES", "LRD RELATED", "PERSONAL STORE", "PR RELATED",
                  "CMS", "FSD","MISC"],
    "MECHANICAL": ['ART/ARME', "CCTV related", "Coaching related (Other)", "MISC", 'Coaching related (Primary)', 'Depot infrastructure (KLBG)', 'Depot infrastructure (KWV)', 'Depot infrastructure (LUR)', 'Depot infrastructure (SUR)', 'Depot infrastructure (WADI', 'HABD related', 'Staff working', 'Wagon related (SUR DIV examined)', 'Wagon related (Other)'],
    "SIGNAL & TELECOM": ["ART/ARME", 'CABLES/EARTHING/KAVACH', 'FIRE ALARM/EXTINGUISHER', 'JOINT INSPECTION (P&C/TC/TRD)', 'LC GATE DEFICIENCIES', 'PANEL/VDU/BI/BPAC/DOCUMENTS', 'PASSENGER AMENITIES/CCTV', 'RELAY ROOM/DL', 'SIGNAL/BOARDS/VEGETATION', 'TRACK CIRCUIT/POINTS', 'WALKIE-TALKIE/COMMUNICATION', 'MISC'],
    "OPTG": ["SWR/CSR/CSL/TWRD", "STATION RECORDS", "STATION DEFICIENCIES", "TRAIN O/P RELATED", "LC GATE DEFICIENCIES", "CIRCULAR/KNOWLEDGE/STAFF", "SIGNAL EXCHANGE", 'WALKIE-TALKIE/PHONE',
             "SM OFFICE DEFICIENCIES/ASSETS", "MISC"],
    "ENGINEERING": ["IOW WORKS (Other)", "IOW WORKS (Safety Related)", "PWI (Track Related)", 'LC GATE DEFICIENCIES', 'P&C', 'WORKSITE', 'Trespass/CRO'],
    "COMMERCIAL": ["REQUIREMENT/ASSETS", "CLEANLINESS/COAL BAGS", "PASSENGER AMENITIES", "STAFF (RAILWAY/CONTRACT)", "MISC"],
    "FINANCE": ["MISC"], "MEDICAL": ["MISC"], "STORE": ["MISC"], "GSU": ["IOW WORKS (Other)", "IOW WORKS (Safety Related)"]
}

INSPECTION_BY_LIST = [""] + ["HQ OFFICER CCE/CR", 'DRM/SUR', 'ADRM', 'Sr.DSO', 'Sr.DOM', 'Sr.DEN/S', 'Sr.DEN/C', 'Sr.DEN/Co', 'Sr.DSTE',
                              'Sr.DEE/TRD', 'Sr.DEE/G', 'Sr.DEE/TRO', 'Sr.DME', 'Sr.DCM', 'Sr.DPO', 'Sr.DFM', 'Sr.DMM', 'DSC',
                              'DME', 'DEE/TRD', 'DFM', 'DSTE/HQ', 'DSTE/KLBG', 'ADEN/T/SUR', 'ADEN/W/SUR', 'ADEN/KWV',
                              'ADEN/PVR', 'ADEN/LUR', 'ADEN/KLBG', 'ADSTE/SUR', 'ADSTE/I/KWV', 'ADSTE/II/KWV', 'ADSTE/KLBG', 'DSTE/SUR'
                              'ADME/SUR', 'AOM/GD', 'AOM/GEN', 'ACM/Cog', 'ACM/TC', 'ACM/GD', 'APO/GEN', 'APO/WEL',
                              'ADFM/I', 'ADFMII', 'ASC', 'ADSO/SUR', "ADME/WADI", 'DEN/TRACK', 'SAFETY TEAM']

ACTION_BY_LIST = [""] + ['DRM/SUR', 'ADRM', 'Sr.DSO', 'Sr.DOM', 'Sr.DEN/S', 'Sr.DEN/C', 'Sr.DEN/Co', 'Sr.DSTE',
                         'Sr.DEE/TRD', 'Sr.DEE/G', 'Sr.DEE/TRO', 'Sr.DME', 'Sr.DCM', 'Sr.DPO', 'Sr.DFM', 'Sr.DMM', 'DSC', 'CMS', 'ADEN/TM/SUR', 'DEN/TRACK', 'ADEN/GSU']

VALID_INSPECTIONS = [
    "FOOTPLATE INSPECTION", "STATION INSPECTION", "LC GATE INSPECTION",
    "COACHING DEPOT", "ON TRAIN", "SURPRISE/AMBUSH INSPECTION", "WORKSITE INSPECTION", "OTHER (UNUSUAL)",
]

# =========================================================================
# HELPERS — data / classification
# =========================================================================
def normalize_str(text):
    if not isinstance(text, str):
        return ""
    return re.sub(r'\s+', ' ', text.lower()).strip()


def classify_feedback(feedback, user_remark=""):
    if isinstance(feedback, str) and feedback.strip() == "`":
        return ""

    def _classify(text_normalized):
        if not text_normalized:
            return None
        date_found = bool(re.search(r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b', text_normalized))
        resolved_kw = [
            "attended", "solved", "done", "completed", "confirmed by", "message given",
            "tdc work completed", "replaced", "msg given", "msg sent", "counseled", "info shared",
            "communicated", "sent successfully", "counselled", "gate will be closed soon",
            "attending at the time", "handled", "resolved", "action taken", "spoken to", "warned",
            "counselling", "hubli", "working normal", "met", "discussion held", "report sent",
            "notified", "explained", "nil", "na", "tlc", "work completed", "acknowledged", "visited",
            "briefed", "guided", "handover", "working properly", "checked found working", "supply restored",
            "this is not a deficiency.", "this is not a deficiency", "not a deficiency", "this is observation",
            "it is observation", "updated by", "adv to", "counselled the staff", "complied", "checked and found",
            "maintained", "for needful action", "provided at", "in working condition", "is working",
            "found working", "equipment is working", "item is working", "as per plan", "putright", "put right",
            'attend dt', 'attend dt.', "operational feasibility", "will be provided",
            "will be supplied shortly", "advised to ubl", "updated"
        ]
        pending_kw = [
            "work is going on", "tdc given", "target date", "expected by", "likely by", "planned by",
            "will be", "needful", "to be", "pending", "not done", "awaiting", "waiting", "yet to", "next time",
            "follow up", "tdc.", "tdc", "t d c", "will attend", "will be attended", "scheduled", "reminder",
            "to inform", "to counsel", "to submit", "to do", "to replace", "prior", "remains", "still",
            "under process", "not yet", "to be done", "will ensure", "during next", "action will be taken",
            'noted please tdc', "will be supplied shortly", "not available", "not updated", "progress",
            "under progress", "to arrange", "awaited", "material awaited", "approval awaited", "to procure",
            "yet pending", "incomplete", "tentative", "ongoing", "in progress", "being done", "arranging",
            "waiting for", "subject to", "awaiting approval", "awaiting material", "awaiting confirmation",
            "next schedule", "planned for", "will arrange", "proposed date", "to complete", "to be completed",
            "likely completion", "expected completion", "not received", "awaiting response"
        ]
        if "tdc" in text_normalized and any(k in text_normalized for k in resolved_kw):
            return "Resolved"
        if any(k in text_normalized for k in pending_kw):
            return "Pending"
        if date_found:
            return "Pending" if "tdc" in text_normalized else "Resolved"
        if any(k in text_normalized for k in resolved_kw):
            return "Resolved"
        return None

    fb = normalize_str(feedback)
    rm = normalize_str(user_remark)
    m = re.findall(r"[!#]", f"{fb} {rm}".strip())
    if m:
        return "Resolved" if m[-1] == "#" else "Pending"
    a = _classify(fb)
    b = _classify(rm)
    if a == "Resolved" or b == "Resolved":
        return "Resolved"
    if a == "Pending" or b == "Pending":
        return "Pending"
    return "Pending"


def get_status(feedback, remark):
    return classify_feedback(feedback, remark)


def color_text_status(status):
    return "🔴 Pending" if status == "Pending" else ("🟢 Resolved" if status == "Resolved" else status)


def strip_status_emoji(status_text):
    """Undo color_text_status() so exports can match on plain 'Pending'/'Resolved'."""
    return re.sub(r"^[^\w]*", "", str(status_text)).strip()


def expand_locations(selected_locations):
    """Expand any selected footplate routes into their constituent sub-locations."""
    expanded = set(selected_locations)
    for loc in selected_locations:
        if loc in FOOTPLATE_ROUTE_HIERARCHY:
            expanded.update(FOOTPLATE_ROUTE_HIERARCHY[loc])
    return expanded


# =========================================================================
# HELPERS — Google Sheet update
# =========================================================================
def update_feedback_column(edited_df):
    header = sheet.row_values(1)

    def col_idx(name):
        try:
            return header.index(name) + 1
        except ValueError:
            st.warning(f"Column '{name}' not found in sheet header.")
            return None

    feedback_col = col_idx("Feedback")
    remark_col = col_idx("User Feedback/Remark")
    head_col = col_idx("Head")
    action_col = col_idx("Action By")
    subhead_col = col_idx("Sub Head")
    timestamp_col = col_idx(TIMESTAMP_COL_NAME)

    required = (feedback_col, remark_col, head_col, action_col, subhead_col, timestamp_col)
    if None in required:
        st.error("Cannot update: one or more required columns missing in Google Sheet.")
        return

    updates = []
    ist = pytz.timezone('Asia/Kolkata')

    for _, row in edited_df.iterrows():
        r = int(row["_sheet_row"])
        def a1(c):
            return gspread.utils.rowcol_to_a1(r, c)

        fv = row.get("Feedback", "") or ""
        rv = row.get("User Feedback/Remark", "") or ""
        hv = row.get("Head", "") or ""
        av = row.get("Action By", "") or ""
        sv = row.get("Sub Head", "") or ""

        timestamp_value = ""
        if fv.strip():
            now_ist = datetime.now(ist)
            timestamp_value = now_ist.strftime("%d-%m-%Y %H:%M:%S IST")

        updates.extend([
            {"range": a1(feedback_col), "values": [[fv]]},
            {"range": a1(remark_col), "values": [[rv]]},
            {"range": a1(head_col), "values": [[hv]]},
            {"range": a1(action_col), "values": [[av]]},
            {"range": a1(subhead_col), "values": [[sv]]},
            {"range": a1(timestamp_col), "values": [[timestamp_value]]},
        ])

        s = st.session_state.df
        mask = s["_sheet_row"] == r
        s.loc[mask, "Feedback"] = fv
        s.loc[mask, "User Feedback/Remark"] = rv
        s.loc[mask, "Head"] = hv
        s.loc[mask, "Action By"] = av
        s.loc[mask, "Sub Head"] = sv
        s.loc[mask, TIMESTAMP_COL_NAME] = timestamp_value

    if updates:
        try:
            sheet.spreadsheet.values_batch_update({
                "valueInputOption": "USER_ENTERED",
                "data": updates
            })
            st.success(f"Updated {len(updates) // 6} record(s) including timestamps.")
        except Exception as e:
            st.error(f"Google Sheets update failed: {str(e)}")


# =========================================================================
# HELPERS — filters
# =========================================================================
def apply_common_filters(df, prefix=""):
    default_to_date = date.today()
    default_from_date = default_to_date - timedelta(days=2)

    with st.expander("🔍 Apply Additional Filters", expanded=True):
        c1, c2 = st.columns(2)
        c1.multiselect(
            "Inspection By", INSPECTION_BY_LIST[1:],
            default=st.session_state.get(prefix + "insp", []),
            key=prefix + "insp"
        )
        c2.multiselect(
            "Action By", ACTION_BY_LIST[1:],
            default=st.session_state.get(prefix + "action", []),
            key=prefix + "action"
        )

        d1, d2 = st.columns(2)
        d1.date_input(
            "📅 From Date",
            value=st.session_state.get(prefix + "from_date", default_from_date),
            key=prefix + "from_date"
        )
        d2.date_input(
            "📅 To Date",
            value=st.session_state.get(prefix + "to_date", default_to_date),
            key=prefix + "to_date"
        )

    out = df.copy()

    if st.session_state.get(prefix + "insp"):
        sel = st.session_state[prefix + "insp"]
        out = out[out["Inspection By"].apply(
            lambda x: any(s.strip() in str(x).split(",") for s in sel)
        )]

    if st.session_state.get(prefix + "action"):
        sel = st.session_state[prefix + "action"]
        out = out[out["Action By"].apply(
            lambda x: any(s.strip() in str(x).split(",") for s in sel)
        )]

    if st.session_state.get(prefix + "from_date") and st.session_state.get(prefix + "to_date"):
        from_date = st.session_state[prefix + "from_date"]
        to_date = st.session_state[prefix + "to_date"]

        if from_date > to_date:
            st.warning("From Date cannot be after To Date. Adjusting filter.")
            from_date, to_date = to_date, from_date

        out = out[
            (out["Date of Inspection"] >= pd.to_datetime(from_date)) &
            (out["Date of Inspection"] <= pd.to_datetime(to_date))
        ]

    return out


# =========================================================================
# HELPERS — Excel export (shared by both download buttons)
# =========================================================================
def build_excel_export(export_df, sheet_name):
    """Build a styled Excel workbook (bytes buffer) from an export-ready DataFrame.
    Expects a 'Date of Inspection' column (date) and a plain-text 'Status' column
    ('Pending' / 'Resolved') with no emoji prefix.
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Apply border + wrap alignment to every cell first
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = thin_border

        # Then layer the date format on top WITHOUT touching border/alignment
        if "Date of Inspection" in export_df.columns:
            date_col_idx = export_df.columns.get_loc("Date of Inspection") + 1
            for row in ws.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx,
                                     max_row=len(export_df) + 1):
                for cell in row:
                    cell.number_format = "DD-MM-YYYY"   # <-- direct format, not cell.style

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = (max_length + 2) if max_length < 50 else 50

        if "Status" in export_df.columns:
            status_col_idx = export_df.columns.get_loc("Status") + 1
            for row in ws.iter_rows(min_row=2, min_col=status_col_idx, max_col=status_col_idx,
                                     max_row=len(export_df) + 1):
                for cell in row:
                    val = strip_status_emoji(cell.value).lower()
                    if val == "pending":
                        cell.font = Font(color="FF0000")  # Red
                    elif val == "resolved":
                        cell.font = Font(color="008000")  # Green

    buf.seek(0)
    return buf


# =========================================================================
# HELPERS — pie-chart breakdown (shared by Head / Sub Head distributions)
# =========================================================================
def render_pie_breakdown(df, group_col, chart_title, caption_parts, threshold=0.02):
    """Pie chart with leader lines + labels, and a counts table on the right.
    Labels are evenly spaced top-to-bottom on each side so they never overlap,
    and the figure/table scale with the number of categories.
    """
    work = df.copy()
    work[group_col] = work[group_col].fillna("").astype(str).str.strip()
    work.loc[work[group_col] == "", group_col] = "(Blank)"

    summary = (
        work.groupby(group_col, sort=False)[group_col]
        .count()
        .reset_index(name="Count")
        .sort_values(by="Count", ascending=False)
        .reset_index(drop=True)
    )
    if summary.empty:
        return

    total = int(summary["Count"].sum())
    summary["Percent"] = summary["Count"] / total

    major = summary[summary["Percent"] >= threshold][[group_col, "Count"]].copy()
    minor = summary[summary["Percent"] < threshold]
    if not minor.empty:
        major = pd.concat(
            [major, pd.DataFrame([{group_col: "Others", "Count": int(minor["Count"].sum())}])],
            ignore_index=True,
        )
    major = major.sort_values("Count", ascending=False).reset_index(drop=True)
    n_slices = len(major)

    base_colors = [
        "#4E79A7", "#F28E2B", "#E15759", "#76B7B2", "#59A14F",
        "#EDC948", "#B07AA1", "#FF9DA7", "#9C755F", "#BAB0AC",
        "#1F77B4", "#FF7F0E", "#2CA02C", "#D62728", "#9467BD",
        "#8C564B", "#E377C2", "#7F7F7F", "#BCBD22", "#17BECF",
    ]
    colors = [base_colors[i % len(base_colors)] for i in range(n_slices)]

    table_rows = [[str(r[group_col]), int(r["Count"])] for _, r in summary.iterrows()]
    table_rows.append(["TOTAL", total])
    n_table_rows = len(table_rows)

    # ---- Dynamic sizing so nothing gets cramped as category count grows ----
    fig_height = max(5.5, 0.34 * n_table_rows + 2.0, 0.5 * n_slices + 2.5)
    fig = plt.figure(figsize=(13, fig_height), facecolor="white")
    ax_pie = fig.add_axes([0.03, 0.08, 0.46, 0.84])
    ax_tbl = fig.add_axes([0.58, 0.06, 0.40, 0.88])
    ax_tbl.axis("off")

    wedges, texts, autotexts = ax_pie.pie(
        major["Count"].tolist(),
        colors=colors,
        startangle=90,
        autopct=lambda pct: f"{pct:.1f}%" if pct >= 4 else "",  # hide autopct on tiny slivers to cut clutter
        pctdistance=0.72,
        textprops=dict(color="white", fontsize=8, fontweight="bold"),
        wedgeprops=dict(edgecolor="white", linewidth=1.2),
    )

    # Angle (deg) and unit-circle x/y for each wedge's midpoint
    angles = [(w.theta1 + w.theta2) / 2.0 for w in wedges]
    xs = [np.cos(np.deg2rad(a)) for a in angles]
    ys = [np.sin(np.deg2rad(a)) for a in angles]

    # Split slices into right-side / left-side labels, ordered top-to-bottom
    right_idx = sorted([i for i in range(n_slices) if xs[i] >= 0], key=lambda i: -ys[i])
    left_idx = sorted([i for i in range(n_slices) if xs[i] < 0], key=lambda i: -ys[i])

    # Vertical range for labels scales with how many need to fit on the busiest side
    max_labels_per_side = max(len(right_idx), len(left_idx), 1)
    label_span = max(1.15, 0.24 * (max_labels_per_side - 1))
    font_size = 8 if max_labels_per_side <= 10 else 7

    def place_labels(idx_list, side):
        n = len(idx_list)
        if n == 0:
            return
        label_ys = [0.0] if n == 1 else np.linspace(label_span, -label_span, n)
        lx = 1.42 if side == "right" else -1.42
        ha = "left" if side == "right" else "right"
        rad = 0.05 if side == "right" else -0.05
        for label_y, i in zip(label_ys, idx_list):
            row = major.iloc[i]
            label = f"{row[group_col]} ({int(row['Count'])})"
            ax_pie.annotate(
                label,
                xy=(0.95 * xs[i], 0.95 * ys[i]),
                xytext=(lx, label_y),
                ha=ha, va="center", fontsize=font_size,
                bbox=dict(boxstyle="round,pad=0.25", facecolor="white", edgecolor="#AAAAAA", alpha=0.95),
                arrowprops=dict(arrowstyle="-", color="#777777", lw=0.7,
                                 connectionstyle=f"arc3,rad={rad}"),
            )

    place_labels(right_idx, "right")
    place_labels(left_idx, "left")

    ax_pie.set_xlim(-2.1, 2.1)
    ax_pie.set_ylim(-(label_span + 0.35), label_span + 0.35)
    ax_pie.set_aspect("equal")

    # ---- Table ----
    col_labels = [group_col, "Count"]
    table_font = 9 if n_table_rows <= 20 else 7.5
    row_scale = 1.5 if n_table_rows <= 20 else max(0.9, 22.0 / n_table_rows)

    table = ax_tbl.table(
        cellText=table_rows,
        colLabels=col_labels,
        loc="center",
        cellLoc="left",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(table_font)
    table.scale(1.05, row_scale)

    n_data = len(table_rows) - 1  # exclude TOTAL
    for j in range(2):
        cell = table[(0, j)]
        cell.set_facecolor("#1F4E79")
        cell.set_text_props(color="white", fontweight="bold", fontsize=table_font, ha="center")

    for i in range(1, n_data + 1):
        alt = "#F5F7FA" if i % 2 == 0 else "#FFFFFF"
        table[(i, 0)].set_facecolor(alt)
        table[(i, 0)].set_text_props(ha="left", fontsize=table_font)
        table[(i, 1)].set_facecolor(alt)
        table[(i, 1)].set_text_props(ha="center", fontsize=table_font)

    last = n_data + 1
    for j in range(2):
        table[(last, j)].set_facecolor("#E8F0FE")
        table[(last, j)].set_text_props(fontweight="bold", fontsize=table_font,
                                        ha="center" if j == 1 else "left")

    for i in range(last + 1):
        table[(i, 0)].set_width(0.70)
        table[(i, 1)].set_width(0.30)

    for _k, cell in table.get_celld().items():
        cell.set_edgecolor("#D0D7DE")
        cell.set_linewidth(0.6)

    fig.suptitle(chart_title, fontsize=14, fontweight="bold", y=0.98, color="#1A1A1A")
    fig.text(0.5, 0.015, " | ".join(caption_parts), ha="center", fontsize=7.5, color="#666666")

    buf = BytesIO()
    plt.savefig(buf, format="png", dpi=160, bbox_inches="tight", facecolor="white")
    buf.seek(0)
    plt.close(fig)

    st.image(buf)
    st.download_button(
        f"📥 Download {chart_title} (PNG)",
        data=buf,
        file_name=f"{group_col.lower().replace(' ', '_')}_distribution.png",
        mime="image/png",
        key=f"dl_{group_col}_{chart_title}",
        use_container_width=True,
    )


# =========================================================================
# HEADER
# =========================================================================
st.markdown(
    """
    <div class="saral-header">
        <img src="https://raw.githubusercontent.com/srdsoproject/testing/main/Central%20Railway%20Logo.png"
             class="saral-logo" alt="Central Railway Logo">
        <div class="saral-header-text">
            <h2 class="saral-initiative">An Initiative by <span class="saral-safety">Safety Department</span>, Solapur Division</h2>
            <h1 class="saral-title">📋 S.A.R.A.L</h1>
            <h3 class="saral-subtitle">(Safety Abnormality Report & Action List – Version 1.3)</h3>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

# =========================================================================
# LOAD DATA
# =========================================================================
@st.cache_data(ttl=5)
def load_data():
    REQUIRED_COLS = [
        "Date of Inspection", "Type of Inspection", "Location",
        "Head", "Sub Head", "Deficiencies Noted",
        "Inspection By", "Action By", "Feedback",
        "User Feedback/Remark", TIMESTAMP_COL_NAME
    ]
    try:
        data = sheet.get_all_values()
        if not data or len(data) < 2:
            st.warning("No data found in Google Sheet. Returning empty DataFrame.")
            return pd.DataFrame(columns=REQUIRED_COLS)
        headers = [c.strip() for c in data[0]]
        df = pd.DataFrame(data[1:], columns=headers)
        for col in REQUIRED_COLS:
            if col not in df.columns:
                df[col] = ""
        df["Date of Inspection"] = pd.to_datetime(df["Date of Inspection"], errors="coerce")
        df["Location"] = df["Location"].astype(str).str.strip().str.upper()
        df["_sheet_row"] = df.index + 2
        return df
    except Exception as e:
        st.error(f"❌ Error loading Google Sheet: {str(e)}")
        st.warning("Returning empty DataFrame to prevent crashes.")
        return pd.DataFrame(columns=REQUIRED_COLS)


if st.session_state.df is None:
    st.session_state.df = load_data()

# =========================================================================
# MAIN TABS
# =========================================================================
tabs = st.tabs(["📝 View Records", "📊 Analytics", "📨 Inspections", " 💡Smart Analysis"])

with tabs[0]:
    df = st.session_state.df
    if df is None or df.empty:
        st.warning("No data available. Please check Google Sheets connection or refresh.")
        st.stop()

    for col in ["Type of Inspection", "Location", "Head", "Sub Head", "Deficiencies Noted",
                "Inspection By", "Action By", "Feedback", "User Feedback/Remark"]:
        if col not in df.columns:
            df[col] = ""

    df["Date of Inspection"] = pd.to_datetime(df["Date of Inspection"], errors="coerce")
    df["_original_sheet_index"] = df.index
    df["Status"] = df.apply(lambda r: classify_feedback(r["Feedback"], r.get("User Feedback/Remark", "")), axis=1)

    start_date = df["Date of Inspection"].min() if not df["Date of Inspection"].isna().all() else pd.Timestamp.today()
    end_date = df["Date of Inspection"].max() if not df["Date of Inspection"].isna().all() else pd.Timestamp.today()

    # Primary filters – these will stack on mobile via CSS
    c1, c2 = st.columns(2)
    c1.multiselect("Type of Inspection", VALID_INSPECTIONS, key="view_type_filter")
    c2.multiselect("Location", ALL_LOCATIONS, key="view_location_filter")

    c3, c4 = st.columns(2)
    c3.multiselect("Head", HEAD_LIST[1:], key="view_head_filter")
    sub_opts = sorted({s for h in st.session_state.view_head_filter for s in SUBHEAD_LIST.get(h, [])})
    c4.multiselect("Sub Head", sub_opts, key="view_sub_filter")

    selected_status = st.selectbox("🔘 Status", ["All", "Pending", "Resolved"], key="view_status_filter")

    filtered = df[(df["Date of Inspection"] >= start_date) & (df["Date of Inspection"] <= end_date)]

    if st.session_state.view_type_filter:
        filtered = filtered[filtered["Type of Inspection"].isin(st.session_state.view_type_filter)]

    if st.session_state.view_location_filter:
        filtered = filtered[filtered["Location"].isin(expand_locations(st.session_state.view_location_filter))]

    if st.session_state.view_head_filter:
        filtered = filtered[filtered["Head"].isin(st.session_state.view_head_filter)]

    if st.session_state.view_sub_filter:
        filtered = filtered[filtered["Sub Head"].isin(st.session_state.view_sub_filter)]

    if selected_status != "All":
        filtered = filtered[filtered["Status"] == selected_status]

    filtered = apply_common_filters(filtered, prefix="view_")
    filtered = filtered.apply(lambda x: x.str.replace("\n", " ") if x.dtype == "object" else x)
    filtered = filtered.sort_values("Date of Inspection")

    # Use the user-selected date range from filters (not full dataset min/max)
    sel_from = st.session_state.get("view_from_date")
    sel_to = st.session_state.get("view_to_date")
    if sel_from and sel_to:
        if sel_from > sel_to:
            sel_from, sel_to = sel_to, sel_from
        dr = f"{sel_from.strftime('%d-%m-%Y')} to {sel_to.strftime('%d-%m-%Y')}"
    elif not filtered.empty and not filtered["Date of Inspection"].isna().all():
        # Fallback: actual span of rows currently shown
        dr = (
            f"{filtered['Date of Inspection'].min().strftime('%d-%m-%Y')} to "
            f"{filtered['Date of Inspection'].max().strftime('%d-%m-%Y')}"
        )
    else:
        dr = "N/A"

    # Metrics – 2×2 on mobile, 4 across on desktop
    col_a, col_b, col_c, col_d = st.columns(4)
    pending_count = (filtered["Status"] == "Pending").sum()
    no_response_count = filtered["Feedback"].isna().sum() + (filtered["Feedback"].astype(str).str.strip() == "").sum()
    resolved_count = (filtered["Status"] == "Resolved").sum()
    col_a.metric("🟨 Pending", pending_count)
    col_b.metric("⚠️ No Response", no_response_count)
    col_c.metric("🟩 Resolved", resolved_count)
    col_d.metric("📊 Total Records", len(filtered))

    # ---- Department-wise (Head) breakdown when Location is selected ----
    if st.session_state.view_location_filter and not filtered.empty:
        st.markdown("### 📊 Department-wise Distribution")
        locations = ", ".join(st.session_state.view_location_filter)
        type_display = ", ".join(st.session_state.view_type_filter) if st.session_state.view_type_filter else "All Types"
        render_pie_breakdown(
            filtered, "Head", "Department-wise Breakdown",
            [f"Date Range: {dr}", f"Locations: {locations}", f"Type: {type_display}"]
        )

    # ---- Sub Head breakdown when Head is selected ----
    if st.session_state.view_head_filter and not filtered.empty:
        st.markdown("### Sub Head Distribution")
        heads = ", ".join(st.session_state.view_head_filter)
        type_display = ", ".join(st.session_state.view_type_filter) if st.session_state.view_type_filter else "All Types"
        location_display = ", ".join(st.session_state.view_location_filter) if st.session_state.view_location_filter else "All Locations"
        caption = [f"Date Range: {dr}", f"Department: {heads}", f"Type: {type_display}", f"Location: {location_display}"]
        if st.session_state.view_sub_filter:
            caption.append(f"Sub Head Filter: {', '.join(st.session_state.view_sub_filter)}")
        render_pie_breakdown(filtered, "Sub Head", "Sub Head Breakdown", caption)

    # ---- Export filtered (read-only) records ----
    export_df = filtered[[
        "Date of Inspection", "Type of Inspection", "Location", "Head", "Sub Head",
        "Deficiencies Noted", "Inspection By", "Action By", "Feedback", "User Feedback/Remark",
        "Status", TIMESTAMP_COL_NAME
    ]].copy()
    export_df["Date of Inspection"] = pd.to_datetime(export_df["Date of Inspection"]).dt.date
    st.download_button(
        "📥 Export Filtered Records to Excel",
        data=build_excel_export(export_df, "Filtered Records"),
        file_name="filtered_records.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # ---------- EDITOR ----------
    if not filtered.empty:
        display_cols = [
            "Date of Inspection", "Type of Inspection", "Head", "Sub Head", "Location",
            "Deficiencies Noted", "Inspection By", "Action By", "Feedback",
            "User Feedback/Remark", TIMESTAMP_COL_NAME
        ]
        valid_cols = [col for col in display_cols if col in filtered.columns]
        if not valid_cols:
            st.error("⚠️ No valid columns found in the DataFrame.")
            st.stop()
        if "Deficiencies Noted" not in valid_cols:
            st.error("⚠️ 'Deficiencies Noted' column is required for search functionality.")
            st.stop()

        editable_filtered = filtered.copy()
        if "_original_sheet_index" not in editable_filtered.columns:
            editable_filtered["_original_sheet_index"] = editable_filtered.index
        if "_sheet_row" not in editable_filtered.columns:
            editable_filtered["_sheet_row"] = editable_filtered.index + 2

        editable_df = editable_filtered[valid_cols + ["_original_sheet_index", "_sheet_row"]].copy()

        if "Date of Inspection" in editable_df.columns:
            editable_df["Date of Inspection"] = pd.to_datetime(
                editable_df["Date of Inspection"], errors="coerce"
            ).dt.date

        if "Feedback" in editable_df.columns and "User Feedback/Remark" in editable_df.columns:
            editable_df.insert(
                editable_df.columns.get_loc("User Feedback/Remark") + 1,
                "Status",
                [get_status(r["Feedback"], r["User Feedback/Remark"]) for _, r in editable_df.iterrows()]
            )
            # Keep a plain-text copy for exports (emoji version is display-only in the grid)
            editable_df["_status_plain"] = editable_df["Status"]
            editable_df["Status"] = editable_df["Status"].apply(color_text_status)

        # ---- Global Search ----
        st.markdown("#### 🔍 Search and Filter")
        search_text = st.text_input("Search All Columns (case-insensitive)", "").strip().lower()
        if search_text:
            mask = editable_df[valid_cols].astype(str).apply(
                lambda col: col.str.contains(search_text, case=False, na=False)
            ).any(axis=1)
            editable_df = editable_df[mask].copy()
            st.info(f"Found {len(editable_df)} matching rows after search.")

        # ---- Column filtering ----
        max_cols = st.slider(
            "Max columns to filter on",
            1, len(valid_cols), min(10, len(valid_cols)),
            key="max_cols_filter"
        )
        candidate_columns = valid_cols[:max_cols]
        selected_columns = st.multiselect(
            "Select columns to filter",
            options=candidate_columns,
            key="column_select_filter"
        )

        if selected_columns:
            df_filtered = editable_df.copy()
            for column in selected_columns:
                col_dtype = editable_df[column].dtype
                if isinstance(col_dtype, pd.CategoricalDtype) or col_dtype == "object":
                    unique_vals = sorted(editable_df[column].dropna().unique())
                    selected_vals = st.multiselect(f"Filter {column}", unique_vals, key=f"filter_{column}")
                    if selected_vals:
                        df_filtered = df_filtered[df_filtered[column].isin(selected_vals)]
                elif is_numeric_dtype(editable_df[column]):
                    _min = float(editable_df[column].min())
                    _max = float(editable_df[column].max())
                    step = (_max - _min) / 100 if _max != _min else 1
                    selected_range = st.slider(f"Filter {column}", _min, _max, (_min, _max), step=step, key=f"range_{column}")
                    df_filtered = df_filtered[df_filtered[column].between(selected_range[0], selected_range[1])]
                elif is_datetime64_any_dtype(editable_df[column]):
                    _min = editable_df[column].min()
                    _max = editable_df[column].max()
                    selected_dates = st.date_input(f"Filter {column}", [_min, _max], min_value=_min, max_value=_max, key=f"date_{column}")
                    if len(selected_dates) == 2:
                        df_filtered = df_filtered[df_filtered[column].between(
                            pd.to_datetime(selected_dates[0]), pd.to_datetime(selected_dates[1])
                        )]
                else:
                    case = st.selectbox(f"Case sensitive for {column}?", ["both", "upper", "lower"], key=f"case_{column}")
                    search_term = st.text_input(f"Filter {column}", key=f"search_{column}")
                    if search_term:
                        if case == "upper":
                            df_filtered = df_filtered[df_filtered[column].str.upper().str.contains(search_term.upper(), na=False)]
                        elif case == "lower":
                            df_filtered = df_filtered[df_filtered[column].str.lower().str.contains(search_term.lower(), na=False)]
                        else:
                            df_filtered = df_filtered[df_filtered[column].str.contains(search_term, case=False, na=False)]
            editable_df = df_filtered
            st.info(f"Applied column filters → {len(editable_df)} rows remaining.")

        # ---- AgGrid configuration ----
        grid_display_df = editable_df.drop(columns=["_status_plain"], errors="ignore")
        gb = GridOptionsBuilder.from_dataframe(grid_display_df)
        
        # flex=1 is the default "share" every column gets of the available width;
        # columns below override that share so wide/narrow content is proportioned
        # sensibly, but everything still always fits the screen width.
        gb.configure_default_column(
            editable=False,
            wrapText=True,
            autoHeight=True,
            resizable=True,
            suppressMovable=False,
            flex=1,
            minWidth=90,
        )
        
        col_flex = {
            "Date of Inspection": (1.0, 110),
            "Type of Inspection": (1.2, 140),
            "Head": (1.0, 110),
            "Sub Head": (1.2, 130),
            "Location": (1.0, 100),
            "Deficiencies Noted": (3.0, 260),   # gets the most room, but still bounded
            "Inspection By": (1.3, 140),
            "Action By": (1.3, 140),
            "Feedback": (1.5, 160),
            "User Feedback/Remark": (2.0, 180),
            "Status": (0.8, 100),
            TIMESTAMP_COL_NAME: (1.3, 150),
        }
        for col_name, (flex_val, min_w) in col_flex.items():
            if col_name in grid_display_df.columns:
                gb.configure_column(col_name, flex=flex_val, minWidth=min_w)
        
        if "User Feedback/Remark" in grid_display_df.columns:
            gb.configure_column(
                "User Feedback/Remark",
                editable=True, wrapText=True, autoHeight=True,
                cellEditor="agTextCellEditor", cellEditorPopup=False,
                cellEditorParams={"maxLength": 4000},
                flex=2, minWidth=180,
            )
        
        gb.configure_column("_original_sheet_index", hide=True)
        gb.configure_column("_sheet_row", hide=True)
        gb.configure_grid_options(
            singleClickEdit=True,
            suppressHorizontalScroll=False,
            enableCellTextSelection=False,   # or just delete this line
            ensureDomOrder=True,
        )
        grid_options = gb.build()
        
        st.markdown("#### 🚈 Inspection Details")
        st.caption("Type your compliance in 'User Feedback/Remark' column. Use column headers to sort. Columns auto-fit the screen width — drag a header edge if you'd still like to fine-tune one.")
        grid_response = AgGrid(
            grid_display_df,
            gridOptions=grid_options,
            update_mode=GridUpdateMode.VALUE_CHANGED,
            height=500,
            allow_unsafe_jscode=True,
            fit_columns_on_grid_load=True,
            theme="streamlit",
        )
        edited_df = pd.DataFrame(grid_response["data"])

        # ---- Export edited grid (plain-text Status so colouring matches) ----
        export_cols = [c for c in valid_cols if c not in ["_original_sheet_index", "_sheet_row"]]
        export_edited_df = edited_df[export_cols].copy()
        export_edited_df["Status"] = edited_df["Status"].apply(strip_status_emoji)
        export_edited_df["Date of Inspection"] = pd.to_datetime(export_edited_df["Date of Inspection"]).dt.date

        st.download_button(
            label="📥 Export Edited Records to Excel",
            data=build_excel_export(export_edited_df, "Edited Records"),
            file_name=f"edited_records_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        # ---- Buttons ----
        b1, b2 = st.columns(2)
        submitted = b1.button("✅ Submit Feedback", use_container_width=True)
        refresh_clicked = b2.button("🔄 Refresh Data", use_container_width=True)

        if refresh_clicked:
            with st.spinner("🔄 Refreshing data from Google Sheets..."):
                st.session_state.df = load_data()
            st.success("✅ Data refreshed successfully!")
            st.rerun()

        if submitted:
            if st.session_state.get("feedback_submitting", False):
                st.warning("⏳ Submission already in progress. Please wait...")
            else:
                st.session_state.feedback_submitting = True
                try:
                    with st.spinner("💾 Saving feedback to Google Sheet... Please do not refresh or close the page."):
                        need_cols = {"_original_sheet_index", "User Feedback/Remark"}
                        if not need_cols.issubset(edited_df.columns) or "Feedback" not in editable_filtered.columns:
                            st.error("⚠️ Required columns are missing from the data.")
                        else:
                            orig = editable_filtered.set_index("_original_sheet_index")
                            new_df = edited_df.set_index("_original_sheet_index")

                            old_remarks = orig["User Feedback/Remark"].fillna("").astype(str)
                            new_remarks = new_df["User Feedback/Remark"].fillna("").astype(str)

                            common_ids = new_remarks.index.intersection(old_remarks.index)
                            diff_mask = new_remarks.loc[common_ids] != old_remarks.loc[common_ids]
                            changed_ids = diff_mask[diff_mask].index.tolist()

                            if changed_ids:
                                diffs = new_df.loc[changed_ids].copy()
                                diffs["_sheet_row"] = orig.loc[changed_ids, "_sheet_row"].values

                                for oid in changed_ids:
                                    user_remark = new_df.loc[oid, "User Feedback/Remark"].strip()
                                    if not user_remark:
                                        continue
                                    # Copy the remark into Feedback and clear the remark field
                                    diffs.at[oid, "Feedback"] = user_remark
                                    diffs.at[oid, "User Feedback/Remark"] = ""
                                    st.session_state.df.at[oid, "Feedback"] = user_remark
                                    st.session_state.df.at[oid, "User Feedback/Remark"] = ""

                                update_feedback_column(diffs.reset_index())
                                st.success(f"✅ Successfully updated {len(changed_ids)} record(s)!")
                            else:
                                st.info("ℹ️ No changes detected in the feedback.")
                except Exception as e:
                    st.error(f"❌ Error during submission: {str(e)}")
                finally:
                    st.session_state.feedback_submitting = False
    else:
        st.info("No deficiencies available to update at the moment.")

# =========================================================================
# FOOTER
# =========================================================================
st.markdown("""
<div style="text-align: center; margin: 28px 0;">
  <div class="adaptive-credit">
    <p>
      <strong>Designed & Developed by</strong>
      <span class="highlight">Safety Department</span>,
      <em>Solapur Division</em>
    </p>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("### 📞 Need Help or Correction in Data?")
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    whatsapp_url = "https://wa.me/919022507772?text=Hello%20Safety%20Department%2C%20I%20need%20assistance%20regarding%20S.A.R.A.L%20data."
    st.markdown(
        f"""
        <div style="text-align: center;">
            <a href="{whatsapp_url}" target="_blank" rel="noopener noreferrer">
                <button style="
                    background-color: #25D366; color: white; font-size: 17px; font-weight: bold;
                    padding: 14px 28px; border: none; border-radius: 50px; cursor: pointer;
                    box-shadow: 0 4px 15px rgba(37, 211, 102, 0.4); transition: all 0.3s ease;
                    max-width: 100%;">
                    📱 Contact Us on WhatsApp<br>
                    <small>+91 90225 07772</small>
                </button>
            </a>
            <p style="margin-top: 14px; color: gray; font-size: 13px; line-height: 1.5;">
                For data corrections:<br>
                ✉️ <a href="mailto:sursafetyposition@gmail.com">sursafetyposition@gmail.com</a><br>
                Rly: 55620
            </p>
        </div>
        """,
        unsafe_allow_html=True
    )

# =========================================================================
# ANALYTICS TAB
# =========================================================================
with tabs[1]:
    st.markdown("### Total Deficiencies Trend (Bar + Trend Line)")
    df = st.session_state.df.copy()

    if "Status" not in df.columns:
        df["Status"] = df.apply(
            lambda r: classify_feedback(r["Feedback"], r.get("User Feedback/Remark", "")), axis=1
        )
    df["Status"] = df["Status"].fillna("Pending").replace({"": "Pending", "NA": "Pending"})
    df["Status"] = df["Status"].str.strip().str.upper().map({
        "PENDING": "Pending", "RESOLVED": "Resolved", "CLOSED": "Resolved"
    }).fillna("Pending")

    if df.empty:
        st.info("No data available for analytics.")
    else:
        df["Date of Inspection"] = pd.to_datetime(df["Date of Inspection"], errors="coerce")
        df = df.dropna(subset=["Date of Inspection"])

        min_date = df["Date of Inspection"].min().date()
        max_date = df["Date of Inspection"].max().date()
        start_date, end_date = st.date_input(
            "Select Inspection Date Range",
            value=(min_date, max_date),
            min_value=min_date,
            max_value=max_date
        )
        df = df[
            (df["Date of Inspection"] >= pd.to_datetime(start_date)) &
            (df["Date of Inspection"] <= pd.to_datetime(end_date))
        ].copy()

        def clean_name(text):
            if pd.isna(text):
                return "UNKNOWN"
            s = str(text).strip()
            s = re.sub(r"[\*\-\_\'\"]", "", s)
            s = re.sub(r"\s+", " ", s).strip()
            return s.upper()

        df["Head_clean"] = df["Head"].apply(clean_name)
        dept_map = {
            "ENGINEERING": "ENGINEERING", "GSU": "GSU",
            "ELECT/G": "ELECT/G", "ELECTG": "ELECT/G",
            "ELECT/TRD": "ELECT/TRD", "ELECT/TRO": "ELECT/TRO",
            "OPTG": "OPTG", "OPERATING": "OPTG",
            "SIGNAL & TELECOM": "SIGNAL & TELECOM",
            "MECHANICAL": "MECHANICAL", "COMMERCIAL": "COMMERCIAL",
            "C&W": "C&W", "SECURITY": "SECURITY", "PERSONNEL": "PERSONNEL",
            "MEDICAL": "MEDICAL", "FINANCE": "FINANCE", "STORE": "STORE",
        }
        df["Head_std"] = df["Head_clean"].map(dept_map).fillna("UNKNOWN")

        if "Location" not in df.columns:
            df["Location"] = ""
        df["Location_clean"] = df["Location"].astype(str).apply(clean_name)
        STATIONS_NORM = {clean_name(x) for x in STATION_LIST}
        df["Is_Station"] = df["Location_clean"].isin(STATIONS_NORM)

        all_locations = set(df["Location_clean"].dropna().unique())
        for main_route, subsections in FOOTPLATE_ROUTE_HIERARCHY.items():
            if main_route in all_locations:
                all_locations.update(subsections)
        df = df[df["Location_clean"].isin(all_locations)]

        # ---- Trend chart ----
        trend = df.groupby(pd.Grouper(key="Date of Inspection", freq="MS")).size().reset_index(name="TotalCount")
        trend = trend.dropna(subset=["Date of Inspection"])

        if not trend.empty:
            trend = trend.sort_values("Date of Inspection")
            base = alt.Chart(trend).encode(
                x=alt.X("Date of Inspection:T", title="Month", axis=alt.Axis(format="%b-%Y")),
                y=alt.Y("TotalCount:Q", title="Total Deficiencies")
            )
            bars = base.mark_bar(color="#1f77b4", cornerRadius=3)
            line = base.transform_regression("Date of Inspection", "TotalCount").mark_line(
                color="red", strokeDash=[6, 4], strokeWidth=2.5
            )
            st.altair_chart(bars + line, use_container_width=True)
        else:
            st.info("No data in selected range.")

        # ---- Department summary (overall) ----
        st.markdown("### Department-wise **Total** Deficiencies Logged")
        dept_counts = df.groupby("Head_std").size().reset_index(name="TotalCount") \
            .sort_values("TotalCount", ascending=False)
        total_deficiencies = dept_counts["TotalCount"].sum()
        dept_counts["color"] = "#ff7f0e"
        dept_counts.loc[:2, "color"] = "red"
        for _, row in dept_counts.iterrows():
            st.markdown(f"- **{row['Head_std']}** : **{row['TotalCount']:,}**")
        st.markdown(f"**Grand Total: {total_deficiencies:,}**")

        dept_chart = alt.Chart(dept_counts).mark_bar().encode(
            x=alt.X("TotalCount:Q", title="Total Deficiencies"),
            y=alt.Y("Head_std:N", sort="-x", title="Department"),
            color=alt.Color("color:N", scale=None),
            tooltip=["Head_std", alt.Tooltip("TotalCount", format=",")]
        ).properties(height=max(280, len(dept_counts) * 36))
        st.altair_chart(dept_chart, use_container_width=True)

        top3 = dept_counts.head(3)
        critical_text = ", ".join([f"**{r['Head_std']}** ({r['TotalCount']:,})" for _, r in top3.iterrows()])
        st.markdown(f"**Critical Departments:** {critical_text}")

        # ---- Top 3 stations ----
        st.markdown("### Top 3 Stations having most logged deficiencies")
        station_df = df[df["Is_Station"]].copy()
        if not station_df.empty:
            top3_stations = (
                station_df.groupby("Location_clean")
                .size()
                .reset_index(name="TotalCount")
                .sort_values("TotalCount", ascending=False)
                .head(3)
                .copy()
            )
            top3_stations["Label"] = top3_stations["Location_clean"]
            top3_stations["color"] = "red"
            chart = alt.Chart(top3_stations).mark_bar().encode(
                x=alt.X("TotalCount:Q", title="Total Deficiencies"),
                y=alt.Y("Label:N", sort="-x", title="Station"),
                color=alt.Color("color:N", scale=None),
                tooltip=["Label", alt.Tooltip("TotalCount", format=",")]
            ).properties(height=220)
            st.altair_chart(chart, use_container_width=True)
        else:
            st.info("No station data found in the selected period.")

        # ---- Location filter → department breakdown ----
        st.markdown("### Department wise deficiencies logged")
        all_locations = sorted(all_locations)
        selected_locations = st.multiselect(
            "Select Locations (Stations / Gates / Routes)",
            options=all_locations,
            default=all_locations
        )
        if selected_locations:
            expanded_locations = expand_locations(selected_locations)
            filtered = df[df["Location_clean"].isin(expanded_locations)].copy()

            dept_breakdown = (
                filtered.groupby("Head_std")
                .size()
                .reset_index(name="TotalCount")
                .sort_values("TotalCount", ascending=False)
            )
            status_breakdown = (
                filtered.groupby(["Head_std", "Status"])
                .size()
                .unstack(fill_value=0)
            )
            status_breakdown.columns = [f"{col}Count" for col in status_breakdown.columns]
            status_breakdown = status_breakdown.reset_index()

            summary_df = dept_breakdown.merge(status_breakdown, on="Head_std", how="left")
            summary_df["PendingCount"] = summary_df.get("PendingCount", 0)
            summary_df["ResolvedCount"] = summary_df.get("ResolvedCount", 0)

            bar_chart = alt.Chart(summary_df).mark_bar(color="#1f77b4").encode(
                x=alt.X("TotalCount:Q", title="Total Deficiencies Logged"),
                y=alt.Y("Head_std:N", title="Department", sort="-x"),
                tooltip=[
                    "Head_std",
                    alt.Tooltip("TotalCount", title="Total", format=","),
                    alt.Tooltip("PendingCount", title="Pending", format=","),
                    alt.Tooltip("ResolvedCount", title="Resolved", format=",")
                ]
            ).properties(height=max(280, len(summary_df) * 38))

            text = bar_chart.mark_text(
                align="left", baseline="middle", dx=3, fontWeight="bold", color="black"
            ).encode(text=alt.Text("TotalCount:Q", format=","))

            final_chart = (bar_chart + text).configure_axis(
                labelFontSize=12, titleFontSize=14
            ).configure_title(fontSize=16)
            st.altair_chart(final_chart, use_container_width=True)

            total = summary_df["TotalCount"].sum()
            pending = summary_df["PendingCount"].sum()
            resolved = summary_df["ResolvedCount"].sum()
            st.markdown(
                f"**Total Deficiencies Logged:** {total:,} | "
                f"**Pending:** {pending:,} | "
                f"**Resolved:** {resolved:,}"
            )

            st.markdown("**Department-wise Breakdown:**")
            for _, row in summary_df.iterrows():
                st.markdown(
                    f"- **{row['Head_std']}**: **Total Deficiencies:** {row['TotalCount']:,} | "
                    f"**Pending:** {row['PendingCount']:,} | "
                    f"**Resolved:** {row['ResolvedCount']:,}"
                )

# =========================================================================
# INSPECTIONS TAB — WhatsApp report generator
# =========================================================================
with tabs[2]:
    st.markdown("### 📨 Safety Inspection Report")
    st.caption(
        "Safety Inspections by inspecting officials:"
    )

    refresh_col, _sp = st.columns([1, 5])
    with refresh_col:
        if st.button("🔄 Refresh", key="insp_refresh_btn", use_container_width=True):
            load_inspections_data.clear()
            st.rerun()

    insp_df = load_inspections_data()

    if insp_df.empty:
        st.warning(
            "No data available from the Inspections sheet. Please check the sheet ID/name, "
            "and that it is shared with the service account."
        )
    else:
        valid_dates = insp_df["Date_parsed"].dropna()
        if valid_dates.empty:
            st.warning("No valid inspection dates could be parsed from column G of the Inspections sheet.")
        else:
            min_d = valid_dates.min().date()
            max_d = valid_dates.max().date()

            fc1, fc2 = st.columns(2)
            insp_from_date = fc1.date_input(
                "📅 From Date", value=max_d, min_value=min_d, max_value=max_d, key="insp_from_date"
            )
            insp_to_date = fc2.date_input(
                "📅 To Date", value=max_d, min_value=min_d, max_value=max_d, key="insp_to_date"
            )

            if insp_from_date > insp_to_date:
                st.warning("From Date cannot be after To Date. Swapping them.")
                insp_from_date, insp_to_date = insp_to_date, insp_from_date

            ranged = insp_df[
                (insp_df["Date_parsed"] >= pd.Timestamp(insp_from_date)) &
                (insp_df["Date_parsed"] <= pd.Timestamp(insp_to_date))
            ].copy()

            if ranged.empty:
                st.info("No inspections found in the selected date range.")
            else:
                official_names = sorted(n for n in ranged["Name"].dropna().unique() if str(n).strip())
                locations = sorted(l for l in ranged["Location"].dropna().unique() if str(l).strip())
                designations = sorted(d for d in ranged["InspectionBy"].dropna().unique() if str(d).strip())

                ff1, ff2, ff3 = st.columns(3)
                selected_officials = ff1.multiselect(
                    "👤 Inspecting Official", official_names, key="insp_name_filter"
                )
                selected_locations = ff2.multiselect(
                    "📍 Location", locations, key="insp_location_filter"
                )
                selected_designations = ff3.multiselect(
                    "🎖️ Designation & HQ", designations, key="insp_designation_filter"
                )

                if selected_officials:
                    ranged = ranged[ranged["Name"].isin(selected_officials)]
                if selected_locations:
                    ranged = ranged[ranged["Location"].isin(selected_locations)]
                if selected_designations:
                    ranged = ranged[ranged["InspectionBy"].isin(selected_designations)]

                ranged["Date_str"] = ranged["Date_parsed"].dt.strftime("%Y-%m-%d")
                grouped = ranged.groupby(["Name", "Phone", "Date_str"], sort=True)

                if len(grouped) == 0:
                    st.info("No matching inspections to build a report for.")
                else:
                    st.markdown(f"**{len(grouped)} report(s) found for the selected filters.**")

                    for (g_name, g_phone, g_date), sub in grouped:
                        designation_series = sub["InspectionBy"].dropna()
                        designation_series = designation_series[designation_series.str.strip() != ""]
                        designation_hq = designation_series.iloc[0] if not designation_series.empty else ""

                        # Sub-group into individual numbered inspections (Type + Location)
                        sections = []
                        for (sec_type, sec_location), sec_rows in sub.groupby(["Type", "Location"], sort=False):
                            deficiencies = [
                                (str(row["Deficiency"]).strip(), str(row["ActionBy"]).strip())
                                for _, row in sec_rows.iterrows()
                                if str(row["Deficiency"]).strip()
                            ]
                            if deficiencies:
                                sections.append((sec_type, sec_location, deficiencies))

                        if not sections:
                            continue

                        message = build_whatsapp_message(g_name, g_phone, designation_hq, g_date, sections)
                        n_deficiencies = sum(len(d) for _, _, d in sections)

                        with st.expander(
                            f"📋 {g_name or 'Unknown'} • {g_date} • {len(sections)} inspection(s), "
                            f"{n_deficiencies} deficiencies"
                        ):
                            st.text_area(
                                "Message Preview (tap inside, select all, copy)",
                                value=message,
                                height=420,
                                key=f"insp_msg_{g_name}_{g_phone}_{g_date}",
                            )

                            b1, b2 = st.columns(2)
                            with b1:
                                # No phone number in the link — this opens WhatsApp's own
                                # share screen and lets the user pick the contact/group manually.
                                wa_url = f"https://wa.me/?text={quote(message)}"
                                st.markdown(
                                    f'<a href="{wa_url}" target="_blank" rel="noopener noreferrer">'
                                    f'<button style="width:100%;min-height:44px;background-color:#25D366;'
                                    f'color:white;border:none;border-radius:10px;font-weight:600;">'
                                    f'📤 Share via WhatsApp</button></a>',
                                    unsafe_allow_html=True,
                                )
                            with b2:
                                safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", str(g_name) or "unknown")
                                st.download_button(
                                    "📥 Download as .txt",
                                    data=message.encode("utf-8"),
                                    file_name=f"inspection_report_{safe_name}_{g_date}.txt",
                                    mime="text/plain",
                                    use_container_width=True,
                                    key=f"insp_dl_{g_name}_{g_phone}_{g_date}",
                                )
    def generate_analysis_image(df, department, date_from, date_to,
                                total, resolved, pending, no_response, resolution_rate):
        """Create a single high-quality LANDSCAPE PNG (no overlapping)."""
        # ---------- prepare data (safe against empty / single-month cases) ----------
        if df.empty or "Month" not in df.columns or "Sub Head" not in df.columns:
            fig, ax = plt.subplots(figsize=(16, 9), dpi=150)
            ax.axis("off")
            ax.text(0.5, 0.5, "No data available for the selected period / department",
                    ha="center", va="center", fontsize=14)
            buf = BytesIO()
            fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
            plt.close(fig)
            buf.seek(0)
            return buf

        month_order = sorted(df["Month"].unique())
        month_names = {m: m.strftime("%b-%Y") for m in month_order}

        # Sub-Head table (NO Share)
        sub = (
            df.groupby(["Sub Head", "Month"])
            .size()
            .unstack(fill_value=0)
        )
        for m in month_order:
            if m not in sub.columns:
                sub[m] = 0
        sub = sub.reindex(columns=month_order, fill_value=0)
        sub["Total"] = sub[month_order].sum(axis=1)
        sub = sub.sort_values("Total", ascending=False)
        if isinstance(sub, pd.Series):
            sub = sub.to_frame().T

        # Officer levels
        officer_tables = []
        _dept_cfg = ASSISTANT_OFFICER_LEVEL.get(_normalize_dept(department))
        if _dept_cfg:
            for level in _dept_cfg["levels"]:
                key = level["key"]
                label = level["label"]
                order = level["order"]
                if key not in df.columns:
                    continue
                level_df = df.dropna(subset=[key])
                if level_df.empty:
                    continue
                grouped = (
                    level_df.groupby([key, "Month"])
                    .size()
                    .unstack(fill_value=0)
                    .reindex(order)
                )
                for m in month_order:
                    if m not in grouped.columns:
                        grouped[m] = 0
                grouped = grouped.reindex(columns=month_order, fill_value=0)
                grouped["Total"] = grouped[month_order].sum(axis=1)
                if isinstance(grouped, pd.Series):
                    grouped = grouped.to_frame().T
                officer_tables.append((label, grouped, order))

        n_officer = len(officer_tables)
        n_sub_rows = min(len(sub), 15) if not sub.empty else 1

        # ========== LANDSCAPE figure ==========
        # Wider & shorter → true landscape
        fig_w = 20
        fig_h = 9.5 + 3.2 * n_officer + 0.22 * n_sub_rows
        fig = plt.figure(figsize=(fig_w, fig_h), dpi=150, facecolor="white")

        # Height ratios tuned for landscape (no crowding)
        height_ratios = [0.85, 0.95, 0.12 + 0.22 * n_sub_rows]
        for _ in range(n_officer):
            height_ratios.extend([0.12, 2.4])

        gs = GridSpec(
            3 + n_officer * 2, 1,
            figure=fig,
            height_ratios=height_ratios,
            hspace=0.45          # generous spacing → no overlap
        )

        # ============================================================
        # 1. HEADER (landscape friendly)
        # ============================================================
        ax_header = fig.add_subplot(gs[0])
        ax_header.set_xlim(0, 20)
        ax_header.set_ylim(0, 1.8)
        ax_header.axis("off")

        ax_header.add_patch(mpatches.FancyBboxPatch(
            (0.15, 0.12), 19.7, 1.55,
            boxstyle="round,pad=0.02,rounding_size=0.12",
            facecolor="#0C2F67", edgecolor="none"))

        ax_header.text(0.5, 1.25, "INDIAN RAILWAYS", color="white",
                       fontsize=12, fontweight="bold", va="center")
        ax_header.text(0.5, 0.90, "SOLAPUR DIVISION", color="#A8C5E2",
                       fontsize=9.5, va="center")
        ax_header.text(0.5, 0.58, "CENTRAL RAILWAY", color="#A8C5E2",
                       fontsize=9.5, va="center")

        ax_header.text(10.0, 1.25,
                       f"SAFETY DEFICIENCIES ANALYSIS OF  {department}  DEPARTMENT",
                       color="white", fontsize=14, fontweight="bold",
                       ha="center", va="center")
        ax_header.text(10.0, 0.70,
                       f"FOR THE PERIOD OF  {date_from.strftime('%d %b %Y')}  –  {date_to.strftime('%d %b %Y')}",
                       color="#A8C5E2", fontsize=11, ha="center", va="center")

        ax_header.text(19.5, 1.25, "Source: SARAL", color="white",
                       fontsize=11, fontweight="bold", ha="right", va="center")

        # ============================================================
        # 2. KPI CARDS
        # ============================================================
        ax_kpi = fig.add_subplot(gs[1])
        ax_kpi.set_xlim(0, 20)
        ax_kpi.set_ylim(0, 2.2)
        ax_kpi.axis("off")

        kpi_data = [
            ("TOTAL RECORDS",          f"{total}",               "#1D4FA3", "#E8F0FE", "100% of Total"),
            ("RESOLVED",               f"{resolved}",            "#159447", "#E6F7ED", f"{resolution_rate:.2f}%"),
            ("NO RESPONSE",            f"{no_response}",         "#D91F2D", "#FDE8E8", f"{(no_response/total*100) if total else 0:.2f}%"),
            ("PENDING",                f"{pending}",             "#E58A00", "#FFF4E0", f"{(pending/total*100) if total else 0:.2f}%"),
            ("OVERALL RESOLUTION RATE", f"{resolution_rate:.2f}%", "#7B2D8E", "#F3E8FF", "(Resolved / Total)"),
        ]

        card_w = 3.6
        gap = 0.28
        start_x = 0.4
        for i, (title, value, color, bg, subtxt) in enumerate(kpi_data):
            x = start_x + i * (card_w + gap)
            ax_kpi.add_patch(mpatches.FancyBboxPatch(
                (x, 0.25), card_w, 1.75,
                boxstyle="round,pad=0.02,rounding_size=0.10",
                facecolor=bg, edgecolor="#CCCCCC", linewidth=1.1))
            ax_kpi.add_patch(mpatches.Rectangle(
                (x, 1.80), card_w, 0.20, facecolor=color, edgecolor="none"))
            ax_kpi.text(x + card_w/2, 1.55, title, color=color,
                        fontsize=9, fontweight="bold", ha="center", va="center")
            ax_kpi.text(x + card_w/2, 1.05, value, color=color,
                        fontsize=22, fontweight="bold", ha="center", va="center")
            ax_kpi.text(x + card_w/2, 0.50, subtxt, color="#555555",
                        fontsize=8.5, ha="center", va="center")

        # ============================================================
        # 3. SUB-HEAD TABLE
        # ============================================================
        ax_sub = fig.add_subplot(gs[2])
        ax_sub.axis("off")
        ax_sub.set_title(f"I — CLASSIFICATION SUB HEAD DISTRIBUTION ({department})",
                         loc="left", fontsize=12, fontweight="bold",
                         color="#123A7A", pad=6)

        if sub.empty:
            ax_sub.text(0.5, 0.5, "No Sub-Head data", ha="center", va="center")
        else:
            col_labels = ["Sub Head"] + [month_names[m] for m in month_order] + ["Total"]
            cell_text = []
            for row in sub.head(15).itertuples():
                name = str(row.Index)[:48]
                vals = [str(int(getattr(row, str(m), 0))) for m in month_order]
                total_val = str(int(getattr(row, "Total", 0)))
                cell_text.append([name] + vals + [total_val])

            table = ax_sub.table(
                cellText=cell_text,
                colLabels=col_labels,
                loc="center",
                cellLoc="center"
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.45)
            for (r, c), cell in table.get_celld().items():
                cell.set_edgecolor("#CCCCCC")
                if r == 0:
                    cell.set_facecolor("#123A7A")
                    cell.set_text_props(color="white", fontweight="bold")
                elif r % 2 == 0:
                    cell.set_facecolor("#F7F9FC")

        # ============================================================
        # 4. OFFICER-LEVEL TABLES + DONUTS
        # ============================================================
        for i, (label, grouped, order) in enumerate(officer_tables):
            # Table
            ax_t = fig.add_subplot(gs[3 + i*2])
            ax_t.axis("off")
            ax_t.set_title(f"II — CLASSIFICATION {label} WISE ({department})",
                           loc="left", fontsize=12, fontweight="bold",
                           color="#123A7A", pad=5)

            col_labels = [label] + [month_names[m] for m in month_order] + ["Total"]
            cell_text = []
            for name in order:
                if name in grouped.index:
                    r = grouped.loc[name]
                    cell_text.append(
                        [name] +
                        [str(int(r.get(m, 0))) for m in month_order] +
                        [str(int(r.get("Total", 0)))]
                    )
                else:
                    cell_text.append([name] + ["0"] * len(month_order) + ["0"])

            table = ax_t.table(
                cellText=cell_text,
                colLabels=col_labels,
                loc="center",
                cellLoc="center"
            )
            table.auto_set_font_size(False)
            table.set_fontsize(8)
            table.scale(1.0, 1.5)
            for (r, c), cell in table.get_celld().items():
                cell.set_edgecolor("#CCCCCC")
                if r == 0:
                    cell.set_facecolor("#123A7A")
                    cell.set_text_props(color="white", fontweight="bold")
                elif r % 2 == 0:
                    cell.set_facecolor("#F7F9FC")

            # Donut – placed with enough space for legend
            ax_d = fig.add_subplot(gs[4 + i*2])
            values = [int(grouped.loc[n, "Total"]) if n in grouped.index else 0 for n in order]
            colors_list = _palette(len(order))

            if sum(values) == 0:
                values = [1]
                colors_list = ["#CCCCCC"]

            wedges, _ = ax_d.pie(
                values,
                colors=colors_list,
                startangle=90,
                wedgeprops=dict(width=0.42, edgecolor="white", linewidth=2)
            )
            ax_d.set_title(f"{label} Wise Distribution", fontsize=11, fontweight="bold", pad=4)
            ax_d.text(0, 0, f"TOTAL\n{total}", ha="center", va="center",
                      fontsize=12, fontweight="bold")

            # Legend on the right – no overlap
            ax_d.legend(
                wedges,
                [f"{n}  ({v})" for n, v in zip(order, values)],
                loc="center left",
                bbox_to_anchor=(1.05, 0.5),
                fontsize=8,
                frameon=False
            )

        # ---------- FOOTER ----------
        fig.text(
            0.5, 0.012,
            f"Source: SARAL System  |  Reporting Department: Safety Department, SUR DIVN, CR  |  "
            f"Analysis Type: Deficiency Analysis  |  Period: {date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}  |  "
            f"Department: {department}  |  Data as on: {date.today().strftime('%d %b %Y')}",
            ha="center", va="bottom", fontsize=8, color="white",
            bbox=dict(boxstyle="round,pad=0.35", facecolor="#0C2F67", edgecolor="none")
        )

        # Final layout – prevents any remaining overlap
        plt.tight_layout(rect=[0.015, 0.035, 0.985, 0.985])
        fig.subplots_adjust(hspace=0.45)

        buf = BytesIO()
        fig.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                    facecolor="white", edgecolor="none")
        plt.close(fig)
        buf.seek(0)
        return buf
with tabs[3]:
        #!/usr/bin/env python3
    """
    Indian Railways – Solapur Division
    Unified Safety Deficiencies Dashboard (Streamlit + Google Sheets)
    ===============================================================
    - Single file
    - Google Sheet connectivity (no local xlsx)
    - Logo & train image from GitHub raw links
    - Date range + department selection
    - Detailed dashboards + Sub-Head wise records + Pending records
    
    Run:
        streamlit run app.py
    """
    
    from __future__ import annotations
    
    import io
    import os
    import re
    import sys
    import tempfile
    from datetime import date, datetime
    from pathlib import Path
    from typing import Callable, Dict, List, Optional, Sequence, Set, Tuple
    
    import numpy as np
    import pandas as pd
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch, Rectangle, Circle
    from PIL import Image
    import streamlit as st
    import requests
    
    # ============================================================
    # ★★★  EASY SETTINGS  ★★★
    # ============================================================
    
    # ---------- Google Sheet ----------
    # Public sheet (Anyone with the link can view) – easiest
    GOOGLE_SHEET_ID = st.secrets["google_sheets"]["sheet_id"]          # e.g. 1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms
    GOOGLE_SHEET_GID = st.secrets["google_sheets"]["sheet_name"]                                 # sheet tab gid (usually 0 for first sheet)
    
    # Optional: Service Account (more secure / private sheets)
    # Put credentials.json in same folder and set USE_SERVICE_ACCOUNT = True
    USE_SERVICE_ACCOUNT = False
    SERVICE_ACCOUNT_FILE = st.secrets.get("google_sheets", {})
    
    # ---------- GitHub raw image links ----------
    LOGO_URL = "https://raw.githubusercontent.com/YOUR_USERNAME/YOUR_REPO/main/indian_railways_logo.png"
    TRAIN_URL = "https://raw.githubusercontent.com/YOUR_USERNAME/YOUR_REPO/main/train.png"
    
    # ---------- Output folder ----------
    OUTPUT_FOLDER = Path("DEPARTMENT_DASHBOARDS")
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    
    # ---------- Default report settings ----------
    REPORT_YEAR = 2026
    REPORT_MONTHS = [4, 5, 6, 7]          # will be overwritten by UI
    START_DATE = None
    END_DATE = None
    PERIOD_TITLE = ""
    SECTION_PERIOD = ""
    DATA_AS_ON = "31 JULY 2026"
    
    MONTH_LABELS = {
        1: f"January-{REPORT_YEAR}", 2: f"February-{REPORT_YEAR}", 3: f"March-{REPORT_YEAR}",
        4: f"April-{REPORT_YEAR}", 5: f"May-{REPORT_YEAR}", 6: f"June-{REPORT_YEAR}",
        7: f"July-{REPORT_YEAR}", 8: f"August-{REPORT_YEAR}", 9: f"September-{REPORT_YEAR}",
        10: f"October-{REPORT_YEAR}", 11: f"November-{REPORT_YEAR}", 12: f"December-{REPORT_YEAR}",
    }
    MONTH_SHORT = {
        1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MAY", 6: "JUN",
        7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC",
    }
    
    # ============================================================
    # Colours & helpers
    # ============================================================
    
    NAVY = "#123A7A"
    DARK_NAVY = "#0C2F67"
    GRID = "#D4DDE8"
    LIGHT_BLUE = "#EEF4FA"
    PALE_YELLOW = "#FFF3D4"
    GREEN = "#11833B"
    RED = "#C81E2A"
    ORANGE = "#D97706"
    PURPLE = "#56319A"
    TEXT = "#222222"
    GRAY = "#6B7280"
    
    WIDTH, HEIGHT = 14, 8
    
    def download_image(url: str, suffix: str = ".png") -> str:
        """Download image from URL to temp file and return path."""
        try:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            tmp.write(r.content)
            tmp.close()
            return tmp.name
        except Exception as e:
            st.error(f"Could not download image from {url}: {e}")
            raise
    
    @st.cache_resource
    def get_logo_path() -> str:
        return download_image(LOGO_URL)
    
    @st.cache_resource
    def get_train_path() -> str:
        return download_image(TRAIN_URL)
    
    def _period_title() -> str:
        if PERIOD_TITLE and str(PERIOD_TITLE).strip():
            return str(PERIOD_TITLE).strip()
        names = []
        for m in REPORT_MONTHS:
            label = MONTH_LABELS.get(m, str(m))
            word = label.split("-")[0].upper() if "-" in str(label) else str(label).upper()
            names.append(word)
        if not names:
            return f"FOR THE SELECTED PERIOD {REPORT_YEAR}"
        if len(names) == 1:
            body = names[0]
        elif len(names) == 2:
            body = f"{names[0]} & {names[1]}"
        else:
            body = ", ".join(names[:-1]) + f" & {names[-1]}"
        return f"FOR THE MONTH OF {body} {REPORT_YEAR}"
    
    def _section_period() -> str:
        if SECTION_PERIOD and str(SECTION_PERIOD).strip():
            return str(SECTION_PERIOD).strip()
        if not REPORT_MONTHS:
            return str(REPORT_YEAR)
        first = MONTH_LABELS.get(REPORT_MONTHS[0], str(REPORT_MONTHS[0]))
        last = MONTH_LABELS.get(REPORT_MONTHS[-1], str(REPORT_MONTHS[-1]))
        first_word = str(first).split("-")[0].upper()
        last_label = str(last).upper() if "-" in str(last) else f"{str(last).upper()}-{REPORT_YEAR}"
        if len(REPORT_MONTHS) == 1:
            return last_label
        return f"{first_word} TO {last_label}"
    
    def _month_header_list() -> List[str]:
        return [MONTH_LABELS.get(m, str(m)) for m in REPORT_MONTHS]
    
    def _month_short_list() -> List[str]:
        return [MONTH_SHORT.get(m, str(m)) for m in REPORT_MONTHS]
    
    def draw_box(ax, x, y, w, h, facecolor="white", edgecolor=GRID, radius=0.05):
        ax.add_patch(FancyBboxPatch(
            (x, y), w, h,
            boxstyle=f"round,pad=0.012,rounding_size={radius}",
            facecolor=facecolor, edgecolor=edgecolor, linewidth=0.8,
        ))
    
    def draw_rect(ax, x, y, w, h, color, edgecolor=None):
        ax.add_patch(Rectangle((x, y), w, h, facecolor=color, edgecolor=edgecolor or color, linewidth=0.5))
    
    def add_text(ax, x, y, text, size=8, weight="normal", color=TEXT, ha="left", va="center"):
        ax.text(x, y, str(text), fontsize=size, fontweight=weight, color=color, ha=ha, va=va, family="DejaVu Sans")
    
    def new_canvas():
        fig = plt.figure(figsize=(WIDTH, HEIGHT), dpi=170)
        fig.patch.set_facecolor("white")
        ax = fig.add_axes([0, 0, 1, 1])
        ax.set_xlim(0, WIDTH)
        ax.set_ylim(0, HEIGHT)
        ax.axis("off")
        return fig, ax
    
    def draw_header(ax, title_lines: Sequence[str], subtitle: str = "Source: SARAL"):
        logo_path = get_logo_path()
        train_path = get_train_path()
        logo = Image.open(logo_path).convert("RGBA")
        ax.imshow(logo, extent=[0.08, 0.88, 7.18, 7.94], aspect="auto", zorder=10)
        add_text(ax, 1.02, 7.72, "INDIAN RAILWAYS", 13, "bold", NAVY)
        add_text(ax, 1.02, 7.46, "SOLAPUR DIVISION", 10.5, "bold", NAVY)
        add_text(ax, 1.02, 7.24, "CENTRAL RAILWAY", 10.5, "bold", NAVY)
        y = 7.62
        for i, line in enumerate(title_lines):
            add_text(ax, 7, y - i * 0.23, line, 16 if i < 2 else 13, "bold", NAVY, "center")
        add_text(ax, 7, 6.98, subtitle, 8.5, "bold", NAVY, "center")
        train = Image.open(train_path).convert("RGBA")
        ax.imshow(train, extent=[12.25, 13.88, 7.20, 7.83], aspect="auto", zorder=10)
    
    def draw_kpi_cards(ax, total, resolved, pending, no_response, y=6.20):
        resolution_rate = (resolved / total * 100) if total else 0.0
        cards = [
            ("TOTAL RECORDS", total, "100% of Total", NAVY, "■"),
            ("RESOLVED", resolved, f"{resolution_rate:.2f}%", GREEN, "✓"),
            ("NO RESPONSE", no_response, f"{no_response / total * 100:.2f}%" if total else "0.00%", RED, "..."),
            ("PENDING", pending, f"{pending / total * 100:.2f}%" if total else "0.00%", ORANGE, "P"),
            ("OVERALL RESOLUTION RATE", f"{resolution_rate:.2f}%", "(Resolved / Total)", PURPLE, "↗"),
        ]
        for i, (title, value, sub_t, color, icon) in enumerate(cards):
            x = 0.20 + i * 2.76
            draw_box(ax, x, y, 2.60, 0.65)
            ax.add_patch(Circle((x + 0.35, y + 0.32), 0.19, facecolor=color, edgecolor="white", linewidth=1))
            add_text(ax, x + 0.35, y + 0.32, icon, 15, "bold", "white", "center")
            add_text(ax, x + 0.66, y + 0.44, title, 7.5, "bold", color)
            add_text(ax, x + 0.66, y + 0.23, str(value), 18, "bold", color)
            add_text(ax, x + 0.66, y + 0.06, sub_t, 7.2, "bold", TEXT)
        return resolution_rate
    
    def draw_footer(ax, dept_text: str, data_as_on: Optional[str] = None):
        data_as_on = data_as_on if data_as_on is not None else DATA_AS_ON
        draw_rect(ax, 0, 0, WIDTH, 0.34, DARK_NAVY)
        add_text(ax, 0.22, 0.17, "Source: SARAL System", 8, color="white")
        add_text(ax, 3.25, 0.17, dept_text, 7.5, color="white")
        add_text(ax, 8.55, 0.17, "Analysis Type: Deficiency Analysis", 8, color="white")
        add_text(ax, 11.35, 0.17, f"Data as on: {data_as_on}", 8, color="white")
    
    def subhead_columns(sub_head_width: float = 2.25, month_width: float = 0.72,
                        total_width: float = 0.65, share_width: float = 0.82, with_share: bool = True):
        cols = [("Sub Head", sub_head_width)]
        for m in REPORT_MONTHS:
            cols.append((MONTH_LABELS.get(m, str(m)), month_width))
        cols.append(("Total", total_width))
        if with_share:
            cols.append(("% Share", share_width))
        return cols
    
    def month_pairs():
        return [(MONTH_LABELS.get(m, str(m)), m) for m in REPORT_MONTHS]
    
    def total_row_values(df: pd.DataFrame, with_share: bool = True):
        vals = ["Total"] + [int((df["Month"] == m).sum()) for m in REPORT_MONTHS] + [len(df)]
        if with_share:
            vals.append("100.00%")
        return vals
    
    def sub_row_values(sub_head, row, with_share: bool = True):
        vals = [sub_head] + [int(row.get(m, 0)) for m in REPORT_MONTHS] + [int(row["Total"])]
        if with_share:
            vals.append(f"{row['Share']:.2f}%")
        return vals
    
    def status_counts(df: pd.DataFrame) -> Tuple[int, int, int, int]:
        total = len(df)
        resolved = df["Status"].str.contains("Resolved", case=False, na=False).sum()
        pending = df["Status"].str.contains("Pending", case=False, na=False).sum()
        no_response = df["Status"].str.contains("No Response", case=False, na=False).sum()
        return total, int(resolved), int(pending), int(no_response)
    
    def filter_months(df: pd.DataFrame, months: Optional[Sequence[int]] = None) -> pd.DataFrame:
        months = list(months if months is not None else REPORT_MONTHS)
        df = df.copy()
        if "Date of Inspection" in df.columns:
            df["Date of Inspection"] = pd.to_datetime(df["Date of Inspection"], errors="coerce")
            df["Month"] = df["Date of Inspection"].dt.month
        elif "Month" not in df.columns:
            return df
    
        before = len(df)
        start = globals().get("START_DATE")
        end = globals().get("END_DATE")
        if start or end:
            d = df["Date of Inspection"]
            mask = d.notna()
            if start:
                mask &= d >= pd.to_datetime(start)
            if end:
                end_ts = pd.to_datetime(end) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
                mask &= d <= end_ts
            out = df[mask].copy()
        else:
            out = df[df["Month"].isin(months)].copy()
        return out
    
    def subhead_table(df: pd.DataFrame, months: Optional[Sequence[int]] = None) -> pd.DataFrame:
        months = list(months if months is not None else REPORT_MONTHS)
        sub = df.groupby(["Sub Head", "Month"]).size().unstack(fill_value=0)
        for m in months:
            if m not in sub.columns:
                sub[m] = 0
        sub["Total"] = sub[months].sum(axis=1)
        total = len(df)
        sub["Share"] = (sub["Total"] / total * 100) if total else 0.0
        return sub.sort_values("Total", ascending=False)
    
    def save_fig(fig, path: str) -> None:
        plt.savefig(path, dpi=180, bbox_inches="tight", pad_inches=0.03, facecolor="white")
        plt.close(fig)
    
    # ============================================================
    # Google Sheet loader
    # ============================================================
    
    @st.cache_data(ttl=300)
    def load_google_sheet() -> pd.DataFrame:
        """Load data from Google Sheet (public or service account)."""
        if USE_SERVICE_ACCOUNT:
            try:
                import gspread
                from google.oauth2.service_account import Credentials
                scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
                creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
                gc = gspread.authorize(creds)
                sh = gc.open_by_key(GOOGLE_SHEET_ID)
                ws = sh.get_worksheet(0)
                data = ws.get_all_records()
                df = pd.DataFrame(data)
            except Exception as e:
                st.error(f"Service account load failed: {e}")
                raise
        else:
            # Public sheet CSV export
            url = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=csv&gid={GOOGLE_SHEET_GID}"
            try:
                df = pd.read_csv(url)
            except Exception as e:
                st.error(f"Could not load public Google Sheet. Check Sheet ID & sharing settings.\n{e}")
                raise
    
        # Clean column names
        df.columns = df.columns.astype(str).str.strip().str.replace(r"\s+", " ", regex=True)
        return df
    
    # ============================================================
    # Department generators (kept from original – only key ones shown fully)
    # You can paste remaining generators (elect_trd, mechanical, operating, commercial, elect_tro, snt)
    # exactly as they were – only change is they now receive df instead of reading Excel.
    # ============================================================
    
    def generate_elect_g(df: pd.DataFrame) -> str:
        required = ["Date of Inspection", "Head", "Sub Head", "Location", "Deficiencies Noted", "Status"]
        for c in required:
            if c not in df.columns:
                raise KeyError(f"Required column '{c}' missing. Available: {df.columns.tolist()}")
    
        df = df.copy()
        df["Date of Inspection"] = pd.to_datetime(df["Date of Inspection"], errors="coerce")
        for c in ["Head", "Sub Head", "Location", "Deficiencies Noted", "Status"]:
            df[c] = df[c].fillna("").astype(str).str.strip()
    
        head_clean = df["Head"].str.upper().str.replace(" ", "", regex=False)
        df = df[head_clean.isin(["ELECT/G", "ELECT/G.", "ELECT-G", "ELECTG"])].copy()
        if len(df) == 0:
            raise ValueError("No records found for Head = ELECT/G")
    
        df["Month"] = df["Date of Inspection"].dt.month
        df = filter_months(df)
        if len(df) == 0:
            raise ValueError("No ELECT/G records in selected period")
    
        # ----- Jurisdiction maps (same as original) -----
        SSE_ELECT_KWV = {
            "KWV", "DHS", "KEM", "BLNI", "BTW", "SEI", "PPJ", "WSB", "KEU", "JNTR",
            "BGVN", "MLM", "BRB", "DD", "MLB", "PVR", "SGLA", "DLGN", "JTRD", "SGRE",
            "ARAG", "KVK", "MRJ", "MKPT", "AAG", "WKA", "MA", "WDS", "WSD",
        }
        SSE_ELECT_SUR = {
            "MKPT", "MA", "AAG", "WKA", "WDS", "DUD", "NGS", "BOT", "AKOR", "SUR",
            "JEUR", "PK", "BALE", "MVE", "MO", "TKWD", "HG", "TLT",
        }
        SSE_ELECT_KLBG = {
            "DUD", "KUI", "GDGN", "GUR", "SVG", "BBD", "KLBG", "TJSP", "HQR", "MR",
            "SDB", "SBD", "WADI", "ME",
        }
        SSE_ELECT_WADI = {"SBD", "SDB", "WADI"}
        SSE_ELECT_LUR = {
            "PJR", "LTRR", "YSI", "DKY", "OSA", "HGL", "LUR", "DRSV", "KMRD",
        }
        ELECT_G_ORDER = [
            "SSE/ELECT/KWV", "SSE/ELECT/SUR", "SSE/ELECT/KLBG",
            "SSE/ELECT/WADI", "SSE/ELECT/LUR",
        ]
    
        station_map: Dict[str, str] = {}
        for loc in SSE_ELECT_KWV:
            station_map[loc] = "SSE/ELECT/KWV"
        for loc in SSE_ELECT_SUR:
            station_map.setdefault(loc, "SSE/ELECT/SUR")
        for loc in SSE_ELECT_KLBG:
            station_map.setdefault(loc, "SSE/ELECT/KLBG")
        for loc in SSE_ELECT_WADI:
            station_map.setdefault(loc, "SSE/ELECT/WADI")
        for loc in SSE_ELECT_LUR:
            station_map.setdefault(loc, "SSE/ELECT/LUR")
    
        SECTION_MAP = {
            "KWV-DD": "SSE/ELECT/KWV", "KWV-MLB": "SSE/ELECT/KWV", "KWV-MRJ": "SSE/ELECT/KWV",
            "KWV-DLGN": "SSE/ELECT/KWV", "KWV-SGRE": "SSE/ELECT/KWV", "KWV-SEI": "SSE/ELECT/KWV",
            "KWV-PVR": "SSE/ELECT/KWV", "KWV-DUD": "SSE/ELECT/KWV",
            "KWV-OSA": "SSE/ELECT/LUR", "KWV-LTRR": "SSE/ELECT/LUR",
            "SUR-KWV": "SSE/ELECT/SUR", "KWV-SUR": "SSE/ELECT/SUR", "SUR-DD": "SSE/ELECT/SUR",
            "DD-SUR": "SSE/ELECT/SUR", "SUR-JEUR": "SSE/ELECT/SUR", "JEUR-SUR": "SSE/ELECT/SUR",
            "SUR-NGS": "SSE/ELECT/SUR", "SUR-BOT": "SSE/ELECT/SUR", "BOT-DUD": "SSE/ELECT/SUR",
            "NGS-BOT": "SSE/ELECT/SUR", "DUD-SUR": "SSE/ELECT/SUR", "SUR-PVR": "SSE/ELECT/SUR",
            "SUR-MRJ": "SSE/ELECT/SUR", "SUR-MO": "SSE/ELECT/SUR", "SUR-SDB": "SSE/ELECT/SUR",
            "SUR-KEM": "SSE/ELECT/SUR", "SUR-HG": "SSE/ELECT/SUR",
            "SUR-WADI": "SSE/ELECT/WADI", "WADI-SUR": "SSE/ELECT/WADI",
            "WADI-KLBG": "SSE/ELECT/WADI", "KLBG-WADI": "SSE/ELECT/KLBG",
            "WADI-SDB": "SSE/ELECT/WADI", "WADI-TLT": "SSE/ELECT/WADI", "WADI-KWV": "SSE/ELECT/WADI",
            "KLBG-SUR": "SSE/ELECT/KLBG", "DUD-KLBG": "SSE/ELECT/KLBG", "KLBG-DUD": "SSE/ELECT/KLBG",
            "LUR-KWV": "SSE/ELECT/LUR", "LTRR-KWV": "SSE/ELECT/LUR", "HGL-KWV": "SSE/ELECT/LUR",
        }
    
        LC_ELECT_G_MAPPING = {
            "LC-19A": "SSE/ELECT/SUR", "LC-40": "SSE/ELECT/SUR", "LC-21": "SSE/ELECT/KWV",
            "LC-42": "SSE/ELECT/SUR",
            "LC-2": "SSE/ELECT/LUR", "LC-4": "SSE/ELECT/LUR", "LC-5": "SSE/ELECT/LUR",
            "LC-6": "SSE/ELECT/LUR", "LC-55": "SSE/ELECT/LUR", "LC-59": "SSE/ELECT/LUR",
            "LC-47": "SSE/ELECT/LUR", "LC-39": "SSE/ELECT/LUR", "LC-34": "SSE/ELECT/LUR",
            "LC-10": "SSE/ELECT/LUR", "LC-36": "SSE/ELECT/LUR",
            "LC-22": "SSE/ELECT/KWV", "LC-24": "SSE/ELECT/KWV", "LC-70": "SSE/ELECT/KWV",
            "LC-31": "SSE/ELECT/KWV", "LC-49": "SSE/ELECT/KWV",
            "LC-74": "SSE/ELECT/SUR", "LC-82": "SSE/ELECT/KLBG", "LC-91": "SSE/ELECT/WADI",
            "LC-1": "SSE/ELECT/WADI", "LC-3": "SSE/ELECT/WADI", "LC-61": "SSE/ELECT/SUR",
            "LC-66": "SSE/ELECT/SUR", "LC-60A": "SSE/ELECT/SUR", "LC-60": "SSE/ELECT/SUR",
        }
    
        def normalize_location(value):
            value = str(value).upper().strip()
            return value.replace(" ", "").replace("_", "-").replace("–", "-").replace("—", "-")
    
        def normalize_text(value):
            return str(value).upper().replace("–", "-").replace("—", "-")
    
        ALL_STATIONS = set()
        for s in (SSE_ELECT_KWV, SSE_ELECT_SUR, SSE_ELECT_KLBG, SSE_ELECT_WADI, SSE_ELECT_LUR):
            ALL_STATIONS.update(s)
        ALL_STATIONS_SORTED = sorted(ALL_STATIONS, key=len, reverse=True)
    
        def stations_found_in_text(text):
            text = normalize_text(text)
            found = []
            for station in ALL_STATIONS_SORTED:
                pattern = r"(?<![A-Z0-9])" + re.escape(station) + r"(?![A-Z0-9])"
                if re.search(pattern, text):
                    found.append(station)
            return found
    
        def classify_from_deficiency(deficiency_text):
            text = normalize_text(deficiency_text)
            for section, jurisdiction in SECTION_MAP.items():
                section_pattern = re.escape(section).replace(r"\-", r"\s*[-/]\s*")
                if re.search(section_pattern, text):
                    return jurisdiction, "DEFICIENCY-SECTION"
            found_stations = stations_found_in_text(text)
            if not found_stations:
                return None, "NOT FOUND"
            jurisdictions_found = []
            for station in found_stations:
                j = station_map.get(station)
                if j and j not in jurisdictions_found:
                    jurisdictions_found.append(j)
            if len(jurisdictions_found) == 1:
                return jurisdictions_found[0], "DEFICIENCY-STATION"
            for i in range(len(found_stations)):
                for j in range(i + 1, len(found_stations)):
                    a, b = found_stations[i], found_stations[j]
                    if a + "-" + b in SECTION_MAP:
                        return SECTION_MAP[a + "-" + b], "DEFICIENCY-PAIR"
                    if b + "-" + a in SECTION_MAP:
                        return SECTION_MAP[b + "-" + a], "DEFICIENCY-PAIR"
            return None, "AMBIGUOUS"
    
        def classify_record(location, deficiency):
            location = normalize_location(location)
            if location == "":
                j, m = classify_from_deficiency(deficiency)
                return (j, "DEFICIENCY-" + m) if j else (None, "UNCLASSIFIED")
            if location in SECTION_MAP:
                return SECTION_MAP[location], "LOCATION-SECTION"
            if location in station_map:
                return station_map[location], "LOCATION-STATION"
            if re.match(r"^LC[- ]?\d+[A-Z]*$", location):
                if location in LC_ELECT_G_MAPPING:
                    return LC_ELECT_G_MAPPING[location], "LC-MAP"
                j, m = classify_from_deficiency(deficiency)
                return (j, "LC-" + m) if j else (None, "LC-NOT-FOUND")
            location_clean = location.replace("/", "-")
            if location_clean in SECTION_MAP:
                return SECTION_MAP[location_clean], "LOCATION-SECTION"
            j, m = classify_from_deficiency(deficiency)
            return (j, "DEFICIENCY-" + m) if j else (None, "UNCLASSIFIED")
    
        results = df.apply(lambda row: classify_record(row["Location"], row["Deficiencies Noted"]), axis=1)
        df["ELECT_G"] = results.apply(lambda x: x[0])
        df["Classification_Method"] = results.apply(lambda x: x[1])
        df["ELECT_G"] = df["ELECT_G"].fillna("Unclassified")
        if (df["ELECT_G"] == "Unclassified").any() and "Unclassified" not in ELECT_G_ORDER:
            ELECT_G_ORDER = list(ELECT_G_ORDER) + ["Unclassified"]
    
        total, resolved, pending, no_response = status_counts(df)
        sub = subhead_table(df)
    
        elect_g = (
            df.groupby(["ELECT_G", "Month"]).size().unstack(fill_value=0)
            .reindex(ELECT_G_ORDER, fill_value=0)
        )
        for m in REPORT_MONTHS:
            if m not in elect_g.columns:
                elect_g[m] = 0
        elect_g["Total"] = elect_g[list(REPORT_MONTHS)].sum(axis=1)
        elect_g["Share"] = (elect_g["Total"] / total * 100) if total else 0
    
        # ---- draw ----
        fig, ax = new_canvas()
        draw_header(ax, [
            "SAFETY DEFICIENCIES ANALYSIS OF ELECTRICAL /",
            "GENERAL DEPARTMENT",
            _period_title(),
        ])
        draw_kpi_cards(ax, total, resolved, pending, no_response)
    
        # Sub-head table
        x, y, w, h = 0.15, 3.45, 6.55, 2.48
        draw_box(ax, x, y, w, h)
        draw_rect(ax, x, y + h - 0.28, w, 0.28, NAVY)
        add_text(ax, x + w / 2, y + h - 0.14,
                 f"II - CLASSIFICATION SUB HEAD DISTRIBUTION ({_section_period()})",
                 9.8, "bold", "white", "center")
        columns = subhead_columns(2.25, 0.72, 0.65, 0.82, with_share=True)
        start_x = x + 0.07
        header_y = y + h - 0.56
        row_height = 0.155
        cx = start_x
        for name, cw in columns:
            draw_rect(ax, cx, header_y - row_height / 2, cw, row_height, LIGHT_BLUE, GRID)
            add_text(ax, cx + cw / 2, header_y, name, 6.9, "bold", NAVY, "center")
            cx += cw
        sub_display = sub.head(11)
        for r, (sub_head, row) in enumerate(sub_display.iterrows()):
            row_y = header_y - (r + 1) * row_height
            cx = start_x
            values = sub_row_values(sub_head, row, with_share=True)
            for c, (_, cw) in enumerate(columns):
                draw_rect(ax, cx, row_y - row_height / 2, cw, row_height,
                          PALE_YELLOW if r == len(sub_display) - 1 else "white", GRID)
                add_text(ax, cx + (0.035 if c == 0 else cw / 2), row_y, str(values[c]),
                         6.5, "bold" if c == 0 else "normal", TEXT,
                         "left" if c == 0 else "center")
                cx += cw
        row_y = header_y - (len(sub_display) + 1) * row_height
        cx = start_x
        total_values = total_row_values(df, with_share=True)
        for c, (_, cw) in enumerate(columns):
            draw_rect(ax, cx, row_y - row_height / 2, cw, row_height, PALE_YELLOW, GRID)
            add_text(ax, cx + (0.035 if c == 0 else cw / 2), row_y, str(total_values[c]),
                     6.5, "bold", TEXT, "left" if c == 0 else "center")
            cx += cw
    
        # Bar chart
        x2, y2, w2, h2 = 6.88, 3.45, 6.97, 2.48
        draw_box(ax, x2, y2, w2, h2)
        draw_rect(ax, x2, y2 + h2 - 0.28, w2, 0.28, NAVY)
        add_text(ax, x2 + w2 / 2, y2 + h2 - 0.14,
                 f"SUB HEAD WISE DISTRIBUTION ({_section_period()})", 9.8, "bold", "white", "center")
        plot = sub.head(10)
        left, right = x2 + 2.20, x2 + 5.85
        top, bottom = y2 + h2 - 0.57, y2 + 0.40
        step = (top - bottom) / max(len(plot) - 1, 1)
        maximum = max(plot["Total"].max(), 1) if len(plot) else 1
        for i, (sub_head, row) in enumerate(plot.iterrows()):
            yy = top - i * step
            label = str(sub_head) if len(str(sub_head)) <= 27 else str(sub_head)[:27] + "..."
            add_text(ax, left - 0.08, yy, label, 6.5, "bold", TEXT, "right")
            bar_width = row["Total"] / maximum * (right - left)
            draw_rect(ax, left, yy - 0.055, bar_width, 0.11, NAVY)
            add_text(ax, right + 0.12, yy, str(int(row["Total"])), 7, "bold", TEXT)
    
        # Jurisdiction table
        x3, y3, w3, h3 = 0.15, 0.50, 6.55, 2.62
        draw_box(ax, x3, y3, w3, h3)
        draw_rect(ax, x3, y3 + h3 - 0.28, w3, 0.28, NAVY)
        add_text(ax, x3 + w3 / 2, y3 + h3 - 0.14, "III - CLASSIFICATION SSE/ELECT WISE",
                 9.8, "bold", "white", "center")
        add_text(ax, x3 + w3 / 2, y3 + h3 - 0.43, "ELECT/G JURISDICTION WISE SUMMARY",
                 7.8, "bold", NAVY, "center")
        headers = ["Month"] + ELECT_G_ORDER + ["TOTAL"]
        widths = [0.72, 1.05, 1.05, 1.05, 1.08, 1.05, 0.55]
        start_x = x3 + 0.06
        header_y = y3 + h3 - 0.73
        row_height = 0.34
        cx = start_x
        for header, cw in zip(headers, widths):
            draw_rect(ax, cx, header_y - row_height / 2, cw, row_height, LIGHT_BLUE, GRID)
            add_text(ax, cx + cw / 2, header_y, header, 5.2, "bold", NAVY, "center")
            cx += cw
        for r, (month_name, month_num) in enumerate(month_pairs()):
            row_y = header_y - (r + 1) * row_height
            cx = start_x
            values = [month_name] + [int(elect_g.loc[j, month_num]) for j in ELECT_G_ORDER] + [int(elect_g[month_num].sum())]
            for c, cw in enumerate(widths):
                draw_rect(ax, cx, row_y - row_height / 2, cw, row_height, "white", GRID)
                add_text(ax, cx + cw / 2, row_y, str(values[c]), 6.2, "normal", TEXT, "center")
                cx += cw
        row_y = header_y - (len(REPORT_MONTHS) + 1) * row_height
        cx = start_x
        values = ["TOTAL"] + [int(elect_g.loc[j, "Total"]) for j in ELECT_G_ORDER] + [int(elect_g["Total"].sum())]
        for c, cw in enumerate(widths):
            draw_rect(ax, cx, row_y - row_height / 2, cw, row_height, PALE_YELLOW, GRID)
            add_text(ax, cx + cw / 2, row_y, str(values[c]), 6.3, "bold", TEXT, "center")
            cx += cw
    
        # Donut
        x4, y4, w4, h4 = 6.88, 0.50, 6.97, 2.62
        draw_box(ax, x4, y4, w4, h4)
        draw_rect(ax, x4, y4 + h4 - 0.28, w4, 0.28, NAVY)
        add_text(ax, x4 + w4 / 2, y4 + h4 - 0.14,
                 f"SSE/ELECT WISE DISTRIBUTION ({_section_period()})", 9.8, "bold", "white", "center")
        donut_ax = fig.add_axes([0.515, 0.095, 0.205, 0.245])
        donut_ax.set_aspect("equal")
        donut_values = elect_g["Total"].values
        donut_colors = ["#1D4FA3", "#159447", "#D91F2D", "#E58A00", "#56319A", GRAY]
        while len(donut_colors) < len(donut_values):
            donut_colors.append(GRAY)
        if donut_values.sum() > 0:
            donut_ax.pie(donut_values, startangle=90,
                         wedgeprops={"width": 0.35, "edgecolor": "white", "linewidth": 1.2},
                         colors=donut_colors, normalize=True)
        donut_ax.text(0, 0.07, "TOTAL", ha="center", va="center", fontsize=9, fontweight="bold", color=NAVY)
        donut_ax.text(0, -0.11, str(int(donut_values.sum())), ha="center", va="center",
                      fontsize=14, fontweight="bold", color=NAVY)
        donut_ax.axis("off")
        for i, j in enumerate(ELECT_G_ORDER):
            yy = y4 + h4 - 0.56 - i * 0.40
            value = int(elect_g.loc[j, "Total"])
            percentage = (value / total * 100) if total else 0
            draw_rect(ax, x4 + 3.25, yy - 0.065, 0.13, 0.13, donut_colors[i])
            add_text(ax, x4 + 3.48, yy, j, 6.7, "bold", TEXT)
            add_text(ax, x4 + 6.25, yy, f"{value} ({percentage:.2f}%)", 6.7, "bold", TEXT, "right")
    
        draw_footer(ax, "Reporting Department: Electrical / General , SUR DIVN, CR")
        out = str(OUTPUT_FOLDER / "ELECT_G_Dashboard.png")
        save_fig(fig, out)
        return out
    
    
    # ============================================================
    # Add remaining generators here (engg, elect_trd, mechanical, operating, commercial, elect_tro, snt)
    # They are identical to your original code – just change the first line from
    #     df = pd.read_excel(...)
    # to
    #     # df already passed
    # and remove the Excel reading part.
    # ============================================================
    
    # Placeholder for other generators – copy from your original file and adapt
    def generate_engineering(df: pd.DataFrame, target_den: Optional[str] = "Sr.DEN/C") -> str:
        # Paste the full original generate_engineering body here
        # Only change: remove pd.read_excel and use the passed df
        raise NotImplementedError("Paste original generate_engineering and adapt")
    
    # Similarly for others...
    
    DASHBOARDS: Dict[str, Callable] = {
        "elect_g": generate_elect_g,
        # "engg": lambda df: generate_engineering(df, "Sr.DEN/C"),
        # "engg_s": lambda df: generate_engineering(df, "Sr.DEN/S"),
        # ... add the rest after pasting
    }
    
    # ============================================================
    # Streamlit UI
    # ============================================================
    
    st.set_page_config(
        page_title="Safety Deficiencies Dashboard",
        page_icon="🚂",
        layout="wide",
    )
    
    st.title("Safety Deficiencies Dashboard Generator")
    st.caption("Solapur Division · Central Railway · Google Sheet + GitHub assets")
    
    st.markdown("---")
    
    # Department selector
    DEPARTMENT_OPTIONS = {
        "ELECT / G": "elect_g",
        "Engineering (Sr.DEN/C)": "engg",
        "Engineering (Sr.DEN/S)": "engg_s",
        "Engineering (DEN/TRACK)": "engg_track",
        "Engineering (Full)": "engg_full",
        "ELECT / TRD": "elect_trd",
        "ELECT / TRO": "elect_tro",
        "S&T (ADSTE)": "snt",
        "Mechanical": "mechanical",
        "Operating": "operating",
        "Commercial": "commercial",
    }
    
    col1, col2 = st.columns([2, 1])
    with col1:
        dept_label = st.selectbox("Select department", list(DEPARTMENT_OPTIONS.keys()))
        dept_key = DEPARTMENT_OPTIONS[dept_label]
    with col2:
        mode = st.radio("What to generate", ["Detailed only"], index=0)  # General removed until master_code provided
    
    st.subheader("Report period")
    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("From date", value=date(2026, 4, 1), format="DD/MM/YYYY")
    with c2:
        end_date = st.date_input("To date", value=date(2026, 7, 31), format="DD/MM/YYYY")
    
    if end_date < start_date:
        st.error("To date cannot be before From date.")
        st.stop()
    
    # Compute months
    report_months = []
    y, m = start_date.year, start_date.month
    while (y, m) <= (end_date.year, end_date.month):
        report_months.append(m)
        if m == 12:
            y += 1
            m = 1
        else:
            m += 1
    report_months = list(dict.fromkeys(report_months))
    year = end_date.year
    
    # Update globals
    REPORT_MONTHS = report_months
    REPORT_YEAR = year
    MONTH_LABELS = {m: f"{MONTH_NAMES[m] if 'MONTH_NAMES' in globals() else m}-{year}" for m in range(1, 13)}
    START_DATE = start_date.isoformat()
    END_DATE = end_date.isoformat()
    PERIOD_TITLE = f"FOR THE PERIOD {start_date.strftime('%d %b %Y').upper()} TO {end_date.strftime('%d %b %Y').upper()}"
    SECTION_PERIOD = f"{start_date.strftime('%d %b').upper()} TO {end_date.strftime('%d %b %Y').upper()}"
    DATA_AS_ON = end_date.strftime("%d %B %Y").upper()
    
    st.info(f"**From:** {start_date.strftime('%d %B %Y')}  |  **To:** {end_date.strftime('%d %B %Y')}  |  **Months:** {report_months}")
    
    generate = st.button("Generate Dashboard", type="primary", use_container_width=True)
    
    if generate:
        with st.spinner("Loading Google Sheet & generating dashboard..."):
            try:
                df_raw = load_google_sheet()
                st.success(f"Loaded {len(df_raw)} rows from Google Sheet")
    
                if dept_key not in DASHBOARDS:
                    st.error(f"Dashboard for '{dept_key}' not yet implemented in this unified file. Paste the generator function.")
                else:
                    path = DASHBOARDS[dept_key](df_raw)
                    st.success("Dashboard generated!")
                    st.image(path, use_container_width=True)
                    with open(path, "rb") as f:
                        st.download_button("Download PNG", f.read(), file_name=Path(path).name, mime="image/png")
            except Exception as e:
                st.exception(e)
    
    st.markdown("---")
    st.caption("Data source: Google Sheet · Assets: GitHub · Output: DEPARTMENT_DASHBOARDS/")
